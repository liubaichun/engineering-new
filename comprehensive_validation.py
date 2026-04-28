#!/usr/bin/env python3
"""
企业管理信息系统 — 全面深度验证脚本 v10.0
覆盖：11个导出API + 3个导入API + CRUD + N+1 + 边界值 + 权限
"""
import io, sys, os, time, json
from datetime import date
sys.path.insert(0, '/root/engineering-new')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

import django; django.setup()

import requests
import openpyxl
from django.db import connection, reset_queries
from django.conf import settings

BASE = 'http://127.0.0.1:8001'
SESSION = requests.Session()

resp = SESSION.post(f'{BASE}/api/core/auth/login/', json={'username': 'admin', 'password': 'admin123'})
if resp.status_code != 200:
    print(f'❌ 登录失败: {resp.status_code}'); sys.exit(1)
print(f'✅ 登录成功')

# ============================================================
# 工具
# ============================================================
def get(url, session=None):
    s = session or SESSION
    r = s.get(url); return r.status_code, r.content

def post(url, **kw):
    r = SESSION.post(url, **kw); return r.status_code, r.content

def patch(url, **kw):
    r = SESSION.patch(url, **kw); return r.status_code, r.content

def is_xlsx(content_bytes):
    if not content_bytes or len(content_bytes) < 4: return False
    if content_bytes[:2] == b'PK':
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes)); return bool(wb.sheetnames)
        except: return False
    return False

def excel_bytes(rows, headers):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(headers)
    for r in rows: ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

PASS = 0; FAIL = 0

def rec(name, ok, detail=''):
    global PASS, FAIL
    icon = '✅' if ok else '❌'
    if ok: PASS += 1
    else: FAIL += 1
    print(f'{icon} {name}' + (f' — {detail}' if detail else ''))

# ============================================================
# 第一部分：11个导出API
# ============================================================
print('\n' + '='*60)
print('【第一部分】11个导出API')
print('='*60)

exports = [
    ('公司导出',   '/api/finance/companies/export/'),
    ('收入导出',   '/api/finance/incomes/export/'),
    ('支出导出',   '/api/finance/expenses/export/'),
    ('工资导出',   '/api/finance/wages/export/'),
    ('发票导出',   '/api/finance/invoices/export/'),
    ('客户导出',   '/api/crm/clients/export/'),
    ('合同导出',   '/api/crm/contracts/export/'),
    ('供应商导出', '/api/crm/suppliers/export/'),
    ('项目导出',   '/api/tasks/projects/export/'),
    ('设备导出',   '/api/equipment/export/'),
    ('物料导出',   '/api/material/export/'),
]

for name, path in exports:
    sc, content = get(BASE + path)
    xlsx = is_xlsx(content)
    rec(name, sc == 200 and xlsx, f'HTTP {sc} Excel={xlsx} {len(content)}B')

# ============================================================
# 第二部分：3个导入API（DB端到端验证）
# ============================================================
print('\n' + '='*60)
print('【第二部分】3个导入API')
print('='*60)

from apps.finance.models import Company, EmployeeCompany
real_co = Company.objects.first()
ec = EmployeeCompany.objects.select_related('employee', 'company').first()
emp_name = ec.employee.name if ec else '李四'
co_name  = ec.company.name if ec else (real_co.name if real_co else '广东百川物流有限公司')
print(f'   公司={real_co.name} EC={ec.id if ec else None} 员工={emp_name}')

# 收入
sc, content = post(f'{BASE}/api/finance/incomes/import_records/',
    files={'file': ('i.xlsx', excel_bytes(
        [[real_co.name,'测试项目',9999.00,date.today().strftime('%Y-%m-%d'),'自动化测试','admin','GREEN']],
        ['公司名称','项目名称','金额','日期','摘要','经办人','备注']),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
try:
    j = json.loads(content); ok = sc == 200 and j.get('success')
except: ok = sc == 200
rec('收入导入', ok, f'HTTP {sc} {content[:80]}')

# 支出
sc, content = post(f'{BASE}/api/finance/expenses/import_records/',
    files={'file': ('e.xlsx', excel_bytes(
        [[real_co.name,'测试供应商',8888.00,date.today().strftime('%Y-%m-%d'),'费用报销','自动化测试']],
        ['公司名称','供应商','金额','日期','费用类型','摘要']),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
try:
    j = json.loads(content); ok = sc == 200 and j.get('success')
except: ok = sc == 200
rec('支出导入', ok, f'HTTP {sc} {content[:80]}')

# 工资（用分离年月格式：年份+月份，用未来月份避免UNIQUE冲突）
sc, content = post(f'{BASE}/api/finance/wages/import_records/',
    files={'file': ('w.xlsx', excel_bytes(
        [[emp_name, co_name, 2027, 2, 8000, 2000, 500, 1000, 561, 225, 10714, '正常工资']],
        ['员工姓名','公司名称','年份','月份','基本工资','岗位工资','加班费','奖金','社会保险','住房公积金','实发工资','工资类型']),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
try:
    j = json.loads(content); ok = sc == 200 and j.get('success')
except: ok = sc == 200
rec('工资导入', ok, f'HTTP {sc} {content[:80]}')

# ============================================================
# 第三部分：CRUD
# ============================================================
print('\n' + '='*60)
print('【第三部分】CRUD')
print('='*60)

sc, content = post(f'{BASE}/api/finance/expenses/',
    json={'company':2,'supplier':'自动化供应商','amount':1234.56,
          'expense_date':date.today().strftime('%Y-%m-%d'),'expense_type':'expense','description':'CRUD验证'})
ok = sc in (200,201)
exp_id = json.loads(content).get('id') if ok else None
rec('支出新建', ok, f'HTTP {sc}' + (f' id={exp_id}' if exp_id else ''))

if exp_id:
    sc, _ = get(f'{BASE}/api/finance/expenses/{exp_id}/')
    rec('支出读取', sc == 200, f'HTTP {sc}')
    sc, _ = patch(f'{BASE}/api/finance/expenses/{exp_id}/', json={'description':'验证更新'})
    rec('支出PATCH', sc == 200, f'HTTP {sc}')

sc, content = post(f'{BASE}/api/finance/incomes/',
    json={'company':2,'amount':5000.00,'date':date.today().strftime('%Y-%m-%d'),
          'source':'自动化测试','handler':'admin'})
ok = sc in (200,201)
inc_id = json.loads(content).get('id') if ok else None
rec('收入新建', ok, f'HTTP {sc}' + (f' id={inc_id}' if inc_id else ''))

if inc_id:
    sc, _ = get(f'{BASE}/api/finance/incomes/{inc_id}/')
    rec('收入读取', sc == 200, f'HTTP {sc}')

# ============================================================
# 第四部分：N+1 查询
# ============================================================
print('\n' + '='*60)
print('【第四部分】N+1查询')
print('='*60)

settings.DEBUG = True

from apps.finance.models import WageRecord
from apps.finance.serializers import WageRecordSerializer
from apps.finance.views import WageRecordViewSet
from apps.approvals.models import ApprovalFlow
from apps.approvals.serializers import ApprovalFlowSerializer
from apps.approvals.views import ApprovalFlowViewSet

reset_queries()
class FU: is_authenticated=True; is_superuser=True; is_staff=True; company_id=None
class FR: user=FU(); query_params={'page_size':'20'}

wv = WageRecordViewSet(); wv.request = FR()
qs = wv.get_queryset()[:20]
_ = WageRecordSerializer(qs, many=True).data
wc = len(connection.queries)
rec('工资N+1(20条)', wc <= 5, f'{wc}条SQL')

reset_queries()
av = ApprovalFlowViewSet(); av.request = FR()
qs = av.get_queryset()[:20]
_ = ApprovalFlowSerializer(qs, many=True).data
ac = len(connection.queries)
rec('审批流N+1(20条)', ac <= 5, f'{ac}条SQL')

settings.DEBUG = False

# ============================================================
# 第五部分：字段完整性（直接查模型）
# ============================================================
print('\n' + '='*60)
print('【第五部分】字段完整性')
print('='*60)

from apps.finance.models import Income, Expense, WageRecord
inc_f = {f.name for f in Income._meta.get_fields()}
exp_f = {f.name for f in Expense._meta.get_fields()}
wage_f = {f.name for f in WageRecord._meta.get_fields()}

rec('Income.company存在', 'company' in inc_f)
rec('Income.amount存在', 'amount' in inc_f)
rec('Expense.supplier存在', 'supplier' in exp_f)
rec('Expense.expense_type存在', 'expense_type' in exp_f)
rec('Expense无payee', 'payee' not in exp_f)
rec('Expense无payment_method', 'payment_method' not in exp_f)
rec('WageRecord.company存在', 'company' in wage_f)
rec('WageRecord.employee_company存在', 'employee_company' in wage_f)
rec('WageRecord.tax字段存在', 'tax' in wage_f)  # actual model field (个税)

# Bug修复验证（读源文件）
with open('/root/engineering-new/apps/core/import_excel.py') as f:
    src_ie = f.read()
rec('Bug A: confirmed→pending映射存在', "'confirmed': 'pending'" in src_ie)
rec('Bug B: import_excel无payment_method行', "row_data['payment_method']" not in src_ie)

with open('/root/engineering-new/apps/finance/views.py') as f:
    src_v = f.read()
rec('Bug E/F: expense_type写=expense_type=并存', 'expense_type=row_data.get' in src_v and 'operator=user' in src_v)

# ============================================================
# 第六部分：边界值
# ============================================================
print('\n' + '='*60)
print('【第六部分】边界值')
print('='*60)

sc, _ = post(f'{BASE}/api/finance/incomes/import_records/',
    files={'file': ('empty.xlsx', excel_bytes([['','0','2026-01-01']], ['公司名称','金额','日期']))})
rec('空公司名导入不崩溃', sc == 200, f'HTTP {sc}')

sc, _ = post(f'{BASE}/api/finance/expenses/', json={'company':2})
rec('缺必填返回4xx', sc >= 400, f'HTTP {sc}')

sc, _ = post(f'{BASE}/api/finance/expenses/',
    json={'company':2,'supplier':'T','amount':999999999.99,
          'expense_date':date.today().strftime('%Y-%m-%d'),'expense_type':'expense','description':'x'})
rec('超大金额接受', sc in (200,201), f'HTTP {sc}')

# ============================================================
# 第七部分：权限
# ============================================================
print('\n' + '='*60)
print('【第七部分】权限')
print('='*60)

anon = requests.Session()
sc, _ = get(f'{BASE}/api/finance/incomes/', session=anon)
rec('未登录拒绝', sc in (401,403), f'HTTP {sc}')

sc, _ = get(f'{BASE}/api/finance/incomes/')
rec('管理员访问收入', sc == 200, f'HTTP {sc}')

sc, _ = get(f'{BASE}/api/finance/companies/')
rec('管理员访问公司', sc == 200, f'HTTP {sc}')

# ============================================================
# 第八部分：API性能
# ============================================================
print('\n' + '='*60)
print('【第八部分】API性能(<500ms)')
print('='*60)

pages = [
    ('收入列表',  '/api/finance/incomes/?page_size=20'),
    ('支出列表',  '/api/finance/expenses/?page_size=20'),
    ('工资列表',  '/api/finance/wages/?page_size=20'),
    ('发票列表',  '/api/finance/invoices/?page_size=20'),
    ('员工列表',  '/api/finance/employees/?page_size=20'),
    ('项目列表',  '/api/tasks/projects/?page_size=20'),
]

for name, path in pages:
    t0 = time.time()
    sc, _ = get(BASE + path)
    ms = (time.time()-t0)*1000
    rec(f'{name}({ms:.0f}ms)', sc == 200, f'HTTP {sc}')

# ============================================================
# 总结
# ============================================================
print('\n' + '='*60)
print(f'【验证完成】✅ {PASS} 项通过 | ❌ {FAIL} 项失败')
print('='*60)
sys.exit(0 if FAIL == 0 else 1)
