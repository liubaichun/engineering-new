from django.db.models import F, Q, Sum
from rest_framework import viewsets, filters, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from .models import Invoice
from .serializers import (
    InvoiceSerializer,
)
from .filters import InvoiceFilter
from apps.core.auth import CSRFExemptSessionAuthentication
from apps.core.permissions import RoleRequired
from apps.core.exceptions import api_error, ErrorCode

# 从共享模块导入工具函数
from .views_common import (
    SafePageNumberPagination,
)

# 导入校验用
import io
from openpyxl import load_workbook
from apps.finance.models import Company as FinanceCompany


class InvoiceViewSet(viewsets.ModelViewSet):
    """发票视图集"""

    queryset = Invoice.objects.all().select_related('company', 'contract')
    serializer_class = InvoiceSerializer
    pagination_class = SafePageNumberPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = InvoiceFilter
    search_fields = ['invoice_no', 'remarks']
    ordering_fields = ['issue_date', 'due_date', 'amount', 'created_at']
    ordering = ['-issue_date', '-created_at']
    authentication_classes = [CSRFExemptSessionAuthentication]
    permission_classes = [permissions.IsAuthenticated, RoleRequired]
    action_perms = {
        None: 'finance:invoice:read',
        'create': 'finance:invoice:create',
        'destroy': 'finance:invoice:delete',
        'partial_update': 'finance:invoice:update',
        'import_records': 'finance:invoice:create',
        'cancel': 'finance:invoice:update',
        'mark_paid': 'finance:invoice:update',
        'issue': 'finance:invoice:update',
        'batch_update': 'finance:invoice:update',
        'match_statement': 'finance:invoice:update',
        'unmatch_statement': 'finance:invoice:update',
        'statement_candidates': 'finance:invoice:read',
        'upload_attachment': 'finance:invoice:update',
        'delete_attachment': 'finance:invoice:update',
    }

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return self.queryset.model.objects.none()
        from apps.core.permissions import get_module_companies

        companies = get_module_companies(self.request.user, 'invoice', 'read')
        if companies is None:
            qs = super().get_queryset()
        else:
            qs = super().get_queryset().filter(Q(company_id__in=companies) | Q(company_id__isnull=True))
        return qs.select_related('company', 'project', 'red_invoice_for', 'matched_bank_statement')

    def perform_create(self, serializer):
        user = self.request.user
        if user and user.is_authenticated:
            kwargs = {}
            if hasattr(user, 'company') and user.company_id:
                kwargs['company_id'] = user.company_id
            serializer.save(**kwargs)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """作废发票"""
        invoice = self.get_object()
        if invoice.status == 'paid':
            return api_error(ErrorCode.INVALID_STATE, '已支付的发票不能作废')
        invoice.status = 'cancelled'
        try:
            invoice.save(update_fields=['status'])
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'作废失败：{str(e)}', status_code=500)
        return Response({'status': 'success', 'message': '发票已作废'})

    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        """标记为已支付"""
        invoice = self.get_object()
        if invoice.status == 'cancelled':
            return api_error(ErrorCode.INVALID_STATE, '已作废的发票不能标记为已支付')
        if invoice.status == 'pending':
            return api_error(ErrorCode.INVALID_STATE, '请先开具发票')
        invoice.status = 'paid'
        try:
            invoice.save(update_fields=['status'])
            # 同步更新关联合同的付款计划
            if invoice.contract_id:
                from apps.crm.models import PaymentPlan

                PaymentPlan.objects.filter(
                    contract_id=invoice.contract_id,
                    status__in=['pending', 'partial'],
                ).update(status='paid', paid_date=timezone.now().date())
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'标记支付失败：{str(e)}', status_code=500)
        return Response({'status': 'success', 'message': '发票已标记为已支付'})

    @action(detail=True, methods=['post'])
    def issue(self, request, pk=None):
        """开具发票"""
        invoice = self.get_object()
        if invoice.status != 'pending':
            return api_error(ErrorCode.INVALID_STATE, '只能开具待开票状态的发票')
        invoice.status = 'issued'
        try:
            invoice.save(update_fields=['status'])
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'开票失败：{str(e)}', status_code=500)
        return Response({'status': 'success', 'message': '发票已开具'})

    @action(detail=False, methods=['get'])
    def years(self, request):
        """返回数据库中实际存在的发票年份列表"""
        from django.db.models.functions import ExtractYear

        years = (
            self.get_queryset()
            .annotate(year=ExtractYear('issue_date'))
            .values_list('year', flat=True)
            .distinct()
            .exclude(year=None)
            .order_by('-year')
        )
        return Response({'years': list(years)})

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """发票汇总统计 — 用于前端顶部三个金额指标"""
        queryset = self.get_queryset()

        # 支持 type 和 company_id 过滤（前端切换Tab和公司时调用）
        invoice_type = request.query_params.get('type')
        company_id = request.query_params.get('company_id')
        invoice_status = request.query_params.get('status')
        if invoice_type:
            queryset = queryset.filter(type=invoice_type)
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        if invoice_status:
            queryset = queryset.filter(status=invoice_status)

        date_min = request.query_params.get('issue_date_min')
        date_max = request.query_params.get('issue_date_max')
        if date_min:
            queryset = queryset.filter(issue_date__gte=date_min)
        if date_max:
            queryset = queryset.filter(issue_date__lte=date_max)

        total_count = queryset.count()

        # 含税金额 = amount + tax_amount（合计）
        gross = queryset.aggregate(total=Sum(F('amount') + F('tax_amount')))['total'] or 0
        # 税金 = tax_amount 合计
        total_tax = queryset.aggregate(total=Sum('tax_amount'))['total'] or 0
        # 不含税金额 = amount 合计
        net_amount = queryset.aggregate(total=Sum('amount'))['total'] or 0

        return Response(
            {
                'total_count': total_count,
                'total_amount': float(gross),  # 含税金额
                'total_tax': float(total_tax),  # 税金
                'net_amount': float(net_amount),  # 不含税金额
            }
        )

    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出发票 Excel"""
        from apps.core.export_excel import export_invoices, make_export_response

        queryset = self.get_queryset()
        company_id = request.GET.get('company_id')
        invoice_type = request.GET.get('type')
        date_start = request.GET.get('date_start')
        date_end = request.GET.get('date_end')
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        if invoice_type:
            queryset = queryset.filter(type=invoice_type)
        if date_start:
            queryset = queryset.filter(issue_date__gte=date_start)
        if date_end:
            queryset = queryset.filter(issue_date__lte=date_end)
        records = queryset.select_related('company', 'project')
        buf = export_invoices(list(records))
        return make_export_response(buf, f'发票_{timezone.now().strftime("%Y%m%d")}.xlsx')

    @action(detail=False, methods=['post'])
    def batch_update(self, request):
        """批量更新发票状态"""
        ids = request.data.get('ids', [])
        action = request.data.get('action')  # 'mark_paid' or 'mark_credited'
        if not ids or not isinstance(ids, list):
            return api_error(ErrorCode.VALIDATION_ERROR, '请选择要更新的发票')
        if action not in ('mark_paid', 'mark_credited'):
            return api_error(ErrorCode.VALIDATION_ERROR, '无效的操作类型')
        invoices = self.get_queryset().filter(id__in=ids)
        count = invoices.count()
        if not count:
            return api_error(ErrorCode.NOT_FOUND, '未找到匹配的发票', status_code=404)
        if action == 'mark_paid':
            invoices.update(status='paid')
        elif action == 'mark_credited':
            invoices.update(is_credited=True)
        return Response({'success': True, 'message': f'已更新 {count} 张发票', 'count': count})

    @action(detail=True, methods=['post'])
    def match_statement(self, request, pk=None):
        invoice = self.get_object()
        statement_id = request.data.get('statement_id')
        if not statement_id:
            return api_error(ErrorCode.VALIDATION_ERROR, '缺少 statement_id')
        from apps.finance.models import BankStatement

        try:
            stmt = BankStatement.objects.get(id=statement_id, company_id=invoice.company_id)
        except BankStatement.DoesNotExist:
            return api_error(ErrorCode.NOT_FOUND, '银行流水不存在', status_code=404)

        # 校验方向匹配：收入发票→银行收入，支出发票→银行支出
        inv_type_name = 'income' if invoice.type == 'income' else 'expense'
        if stmt.direction != inv_type_name:
            return api_error(
                ErrorCode.VALIDATION_ERROR,
                f'方向不匹配：{invoice.get_type_display()}不能核销{stmt.get_direction_display()}银行流水',
            )

        invoice.matched_bank_statement = stmt
        invoice.payment_date = stmt.transaction_date
        invoice.status = 'paid'
        try:
            invoice.save(update_fields=['matched_bank_statement', 'payment_date', 'status'])
            # 同步更新银行流水核销状态
            stmt.reconcile_status = 'matched'
            stmt.reconcile_time = timezone.now()
            stmt.save(update_fields=['reconcile_status', 'reconcile_time'])
            # 如果发票关联了合同，尝试同步付款计划
            if invoice.contract_id:
                from apps.crm.models import PaymentPlan

                PaymentPlan.objects.filter(
                    contract_id=invoice.contract_id,
                    status__in=['pending', 'partial'],
                ).update(status='paid', paid_date=stmt.transaction_date)
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'核销失败：{str(e)}', status_code=500)
        return Response({'success': True, 'message': f'已核销：{stmt.counterparty_name} ¥{stmt.amount}'})

    @action(detail=True, methods=['post'])
    def unmatch_statement(self, request, pk=None):
        invoice = self.get_object()
        stmt = invoice.matched_bank_statement
        invoice.matched_bank_statement = None
        try:
            invoice.save(update_fields=['matched_bank_statement'])
            if stmt:
                stmt.reconcile_status = 'unmatched'
                stmt.reconcile_time = None
                stmt.save(update_fields=['reconcile_status', 'reconcile_time'])
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'取消核销失败：{str(e)}', status_code=500)
        return Response({'success': True, 'message': '已取消核销'})

    @action(detail=True, methods=['get'])
    def statement_candidates(self, request, pk=None):
        invoice = self.get_object()
        from apps.finance.models import BankStatement

        candidates = BankStatement.objects.filter(company_id=invoice.company_id, reconcile_status='unmatched').order_by(
            '-transaction_date'
        )[:50]
        from apps.finance.serializers import BankStatementSerializer

        return Response(BankStatementSerializer(candidates, many=True).data)

    @action(detail=True, methods=['post'])
    def upload_attachment(self, request, pk=None):
        invoice = self.get_object()
        file = request.FILES.get('file')
        if not file:
            return api_error(ErrorCode.VALIDATION_ERROR, '请选择文件')
        invoice.attachment = file
        invoice.attachment_name = file.name
        try:
            invoice.save(update_fields=['attachment', 'attachment_name'])
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'上传失败：{str(e)}', status_code=500)
        return Response({'success': True, 'message': f'已上传：{file.name}'})

    @action(detail=True, methods=['post'])
    def delete_attachment(self, request, pk=None):
        invoice = self.get_object()
        invoice.attachment.delete()
        invoice.attachment_name = ''
        try:
            invoice.save(update_fields=['attachment', 'attachment_name'])
        except Exception as e:
            return api_error(ErrorCode.INTERNAL_ERROR, f'删除附件失败：{str(e)}', status_code=500)
        return Response({'success': True, 'message': '附件已删除'})

    @action(detail=False, methods=['post'])
    def import_records(self, request):
        """批量导入发票 Excel（收到/开出的数电发票）"""
        file = request.FILES.get('file')
        if not file:
            return api_error(ErrorCode.VALIDATION_ERROR, '请上传 Excel 文件')

        invoice_type = request.data.get('type')  # 'income' 收到 或 'expense' 开出
        if invoice_type not in ('income', 'expense'):
            return api_error(ErrorCode.VALIDATION_ERROR, '缺少或无效的 type 参数（income/expense）')

        user = request.user
        company_id = request.data.get('company_id')
        if not company_id:
            if hasattr(user, 'company_id') and user.company_id:
                company_id = user.company_id
            company_id = getattr(request, 'auth_company', None) and request.auth_company.id or company_id

        # ─── 发票类型校验：检查Excel中的购方/销方是否与所选入口类型匹配 ──
        if company_id:
            try:
                company_name = FinanceCompany.objects.get(id=company_id).name
                if company_name:
                    file.seek(0)
                    wb = load_workbook(io.BytesIO(file.read()))
                    ws = wb.active
                    # 找表头行
                    header_row = None
                    for row in ws.iter_rows(values_only=True):
                        row_str = [str(v or '') for v in row]
                        if any('购方名称' in s or '销方名称' in s for s in row_str):
                            header_row = row_str
                            break
                    if header_row:
                        try:
                            buyer_idx = next(i for i, h in enumerate(header_row) if '购方名称' in h)
                            seller_idx = next(i for i, h in enumerate(header_row) if '销方名称' in h)
                            # 取第一行数据
                            data_rows = list(ws.iter_rows(min_row=ws._current_row + 1, values_only=True))
                            if data_rows and data_rows[0]:
                                first_buyer = str(data_rows[0][buyer_idx] or '').strip()
                                first_seller = str(data_rows[0][seller_idx] or '').strip()
                                cn = company_name.strip()
                                if invoice_type == 'expense':
                                    # 收到发票：我方是购方，Excel中购方名称应包含公司名
                                    if first_buyer and cn not in first_buyer and first_buyer not in cn:
                                        return api_error(
                                            ErrorCode.VALIDATION_ERROR,
                                            f'检测到Excel中购方名称为「{first_buyer}」，与当前公司「{cn}」不匹配。'
                                            f'您可能在「收到发票」入口上传了开出发票的Excel，请检查后重试。',
                                        )
                                else:
                                    # 开出发票：我方是销方，Excel中销方名称应包含公司名
                                    if first_seller and cn not in first_seller and first_seller not in cn:
                                        return api_error(
                                            ErrorCode.VALIDATION_ERROR,
                                            f'检测到Excel中销方名称为「{first_seller}」，与当前公司「{cn}」不匹配。'
                                            f'您可能在「开出发票」入口上传了收到发票的Excel，请检查后重试。',
                                        )
                        except (StopIteration, IndexError):
                            pass  # 列名找不到就不校验
                    file.seek(0)
            except Exception:
                pass  # 校验失败不阻塞导入

        try:
            from apps.core.import_excel import import_invoice

            result = import_invoice(file, invoice_type=invoice_type, company_id=company_id)
        except Exception as e:
            return api_error(ErrorCode.VALIDATION_ERROR, f'解析失败：{str(e)}')

        if not result.rows:
            # 直接展示具体错误原因，不包裹"未识别到有效发票记录"
            detail = result.errors[0]['message'] if result.errors else '文件解析后无有效数据'
            return Response(
                {
                    'success': False,
                    'message': detail,
                    'errors': result.errors,
                },
                status=400,
            )

        created = 0
        errors = []
        skipped = 0
        # 预查已有发票号避免逐条查询
        existing_nos = set(
            Invoice.objects.filter(
                invoice_no__in=[r['invoice_no'] for r in result.rows if r.get('invoice_no')]
            ).values_list('invoice_no', flat=True)
        )
        for row_data in result.rows:
            inv_no = row_data.get('invoice_no', '')
            if inv_no in existing_nos:
                skipped += 1
                errors.append({'row': 0, 'message': f'发票号 {inv_no} 已存在，跳过'})
                continue
            try:
                # 只取 Invoice 模型已知的字段，避免传入多余字段
                safe_fields = {
                    'invoice_no',
                    'type',
                    'invoice_type',
                    'amount',
                    'tax_rate',
                    'tax_amount',
                    'counterparty',
                    'counterparty_tax_id',
                    'counterparty_bank',
                    'company_id',
                    'issue_date',
                    'status',
                    'is_credited',
                    'credited_period',
                    'due_date',
                    'payment_date',
                    'remarks',
                    'red_invoice_for',
                    'attachment',
                }
                clean = {k: v for k, v in row_data.items() if k in safe_fields}
                obj = Invoice.objects.create(**clean)

                # 红冲发票：自动匹配原始发票
                if float(obj.amount or 0) < 0 and not obj.red_invoice_for:
                    orig = (
                        Invoice.objects.filter(
                            counterparty__exact=obj.counterparty,
                            amount__exact=abs(float(obj.amount)),
                            type=obj.type,
                            company_id=obj.company_id,
                            red_invoice_for__isnull=True,
                        )
                        .exclude(id=obj.id)
                        .order_by('issue_date')
                        .first()
                    )
                    if orig:
                        obj.red_invoice_for = orig
                        obj.status = 'cancelled'
                        obj.save(update_fields=['red_invoice_for', 'status'])

                existing_nos.add(obj.invoice_no)
                created += 1
            except Exception as e:
                errors.append({'row': 0, 'message': f'发票号{inv_no}: {str(e)}'})

        msg_parts = [f'成功 {created} 条']
        if skipped:
            msg_parts.append(f'跳过 {skipped} 条（已存在）')
        if errors:
            msg_parts.append(f'失败 {len(errors)} 条')

        return Response(
            {
                'success': True,
                'message': '导入完成：' + '，'.join(msg_parts),
                'errors': (result.errors + errors)[:20],
                'created': created,
                'skipped': skipped,
            }
        )
