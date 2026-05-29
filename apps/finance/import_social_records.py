"""
社保申报记录导入函数 - 适配深圳社保局「社保费申报明细」格式。
单Sheet，宽表结构：每个险种占3列（缴费工资/费率/应缴费额）× 2（单位/个人）。
"""

import io
import re
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook


def _parse_decimal(value):
    """解析数值，兼容 None/空/带符号/带%的值"""
    if value is None or value == '':
        return None
    s = str(value).replace(',', '').replace('¥', '').replace(' ', '').strip()
    if s == '' or s.lower() in ('n/a', '-', '—'):
        return None
    # 去掉百分号
    if '%' in s:
        try:
            return Decimal(s.replace('%', '')) / 100
        except InvalidOperation:
            return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


# 列结构（0索引）映射表，适配深圳社保局 Excel
# 每个险种占6列：单位缴费工资、单位费率、单位应缴、个人缴费工资、个人费率、个人应缴
INSURANCE_COLUMNS = [
    # 养老
    ('pension_company', 'basic_pension_company', 9, 10, 11),
    ('pension_employee', 'basic_pension_employee', 12, 13, 14),
    # 医疗
    ('medical_company', 'basic_medical_company', 15, 16, 17),
    ('medical_employee', 'basic_medical_employee', 18, 19, 20),
    # 生育
    ('maternity_company', 'maternity_company', 21, 22, 23),
    ('maternity_employee', 'maternity_employee', None, None, None),
    # 工伤
    ('injury_company', 'injury_company', 25, 26, 27),
    ('injury_employee', 'injury_employee', None, None, None),
    # 失业
    ('unemployment_company', 'unemployment_company', 33, 34, 35),
    ('unemployment_employee', 'unemployment_employee', 36, 37, 38),
    # 地方补充养老（单位）
    ('pension_local_company', 'local_pension_company', 39, 40, 41),
]


# 从 Row 1 的险种名称行找到各险种起始列
def _find_insurance_columns(row1_headers):
    """根据 Row 1 的险种名称定位起始列"""
    COLUMNS_MAP = {}
    # 养老保险（单位）从第10列开始
    if len(row1_headers) > 10:
        COLUMNS_MAP['pension_company'] = 9  # 0-indexed
        COLUMNS_MAP['basic_pension_company'] = 10
        COLUMNS_MAP['basic_pension_amount_company'] = 11
    return COLUMNS_MAP


def import_social_records(file_obj, company_id=None):
    """
    导入社保申报记录 Excel。
    返回 dict: {success, message, created, updated, errors}
    """
    wb = load_workbook(io.BytesIO(file_obj.read()), data_only=True)
    ws = wb.active  # 单Sheet，直接用 active

    max_col = ws.max_column
    max_row = ws.max_row

    # --- 第一步：解析 Row 2（列定义：缴费工资/费率/应缴费额）确定每组3列 ---
    # 险种分组结构（0索引列号）：
    #  9-11   基本养老保险（单位）   缴费工资/费率/应缴
    #  12-14  基本养老保险（个人）   缴费工资/费率/应缴
    #  15-17  基本医疗保险（单位）
    #  18-20  基本医疗保险（个人）
    #  21     地方补充医疗（单位）   只有应缴额（占col 21-23，但中间空）
    #  22-23  无实际内容
    #  24-26  基本医疗保险（个人）再确认...
    # 通过实际数据行验证列位置

    # 用 Row 8（小计行）的数据验证列结构
    # Row 8 数据: col[4]='', col[5]=11555.28, col[6]=3402.75, col[7]=8152.53, col[8]='',
    #            col[9]='', col[10]='', col[11]=5172.0 (养老单位应缴小计)
    #            col[13]=2586.0 (养老个人应缴小计)
    #            col[15]=2290.86 (医疗单位)
    #            col[19]=763.62 (医疗个人)
    #            col[21]=190.92 (生育)
    #            col[25]=106.25 (工伤)
    #            col[33]=212.5 (失业单位)
    #            col[35]=53.13 (失业个人)
    #            col[39]=180.0 (地方补充养老)

    records_to_save = []

    # --- 扫描数据行（从 Row 4 开始，跳过表头 Row1-3 和小计行 Row8/10/12）---
    SKIP_LABELS = {'小计', '合计', '在职人员', '退休人员', '家属统筹人员', ''}

    for row_num in range(4, max_row + 1):
        seq_val = _cell_raw(ws, row_num, 0)
        name = _cell_raw(ws, row_num, 1) or ''
        id_card = str(_cell_raw(ws, row_num, 2) or '').strip()
        year_month = str(_cell_raw(ws, row_num, 3) or '').strip()  # e.g. "202601"
        total_receivable = _parse_decimal(_cell_raw(ws, row_num, 5))
        total_employee = _parse_decimal(_cell_raw(ws, row_num, 6))
        total_company = _parse_decimal(_cell_raw(ws, row_num, 7))

        # 跳过小计/合计/分组行/空行
        if name in SKIP_LABELS or (not id_card and not name):
            continue

        # 费款所属期格式化为 YYYY-MM
        if len(year_month) >= 6:
            year_month = f'{year_month[:4]}-{year_month[4:6]}'
        else:
            year_month = None

        # 收集分险种数据（基于深圳社保局实际格式）：
        # 规律：col[n]="费率", col[n+1]="应缴费额" → 金额在 col[n+1]
        # 缴费工资统一在 col[8]，各险种不重复
        data = {
            # 养老单位 col[10]（col[9]是费率16%，所以col[10]=764.0）
            'pension_company_amount': _parse_decimal(_cell_raw(ws, row_num, 10)),
            # 养老个人 col[12]（col[11]是费率8%）
            'pension_employee_amount': _parse_decimal(_cell_raw(ws, row_num, 12)),
            # 医疗单位 col[14]（col[13]是费率6%）
            'medical_company_amount': _parse_decimal(_cell_raw(ws, row_num, 14)),
            # 医疗个人 col[18]（col[17]是费率2%）
            'medical_employee_amount': _parse_decimal(_cell_raw(ws, row_num, 18)),
            # 生育 col[20]（col[19]是费率0.5%）
            'maternity_company_amount': _parse_decimal(_cell_raw(ws, row_num, 20)),
            # 工伤 col[26]（col[25]是费率0.4%）
            'injury_company_amount': _parse_decimal(_cell_raw(ws, row_num, 26)),
            # 失业单位 col[34]（col[33]是费率0.8%）
            'unemployment_company_amount': _parse_decimal(_cell_raw(ws, row_num, 34)),
            # 失业个人 col[36]（col[35]是费率0.2%）
            'unemployment_employee_amount': _parse_decimal(_cell_raw(ws, row_num, 36)),
            # 地补养老单位 col[38]（col[37]是费率1%）
            'local_pension_company_amount': _parse_decimal(_cell_raw(ws, row_num, 38)),
            # 缴费工资（用于参考/稽核）
            'pension_wage': _parse_decimal(_cell_raw(ws, row_num, 8)),
            'medical_wage': _parse_decimal(_cell_raw(ws, row_num, 8)),
        }

        records_to_save.append(
            {
                'seq': seq_val,
                'name': name,
                'id_card': id_card,
                'year_month': year_month,
                'total_receivable': total_receivable,
                'total_employee': total_employee,
                'total_company': total_company,
                **data,
            }
        )

    if not records_to_save:
        return {'success': False, 'message': '未能从文件中识别到有效数据行，请检查文件格式'}

    # --- 第二步：写入 SocialRecord ---
    from apps.finance.models import SocialRecord, Employee

    # 从 Sheet 名解析公司名（格式：「深圳市XXX_社保费申报明细_20260128」）
    actual_company_id = company_id
    if not actual_company_id:
        from apps.finance.models import Company

        sheet_name = ws.title or ''
        m = re.match(r'^(.+?)_社保费申报明细', sheet_name)
        if m:
            company_name = m.group(1).strip()
            # 精确匹配
            comp = Company.objects.filter(name=company_name).first()
            if comp:
                actual_company_id = comp.id
            else:
                # 模糊匹配
                comp = Company.objects.filter(name__contains=company_name).first()
                if comp:
                    actual_company_id = comp.id
        if not actual_company_id:
            return {'success': False, 'message': f'未能从文件Sheet名「{sheet_name}」识别到公司，请手动选择公司后重试'}

    created = 0
    updated = 0
    errors = []
    skipped_dirty = []  # 跳过的脏数据（仅工伤无其他险）

    for rec in records_to_save:
        if not rec['id_card'] or len(rec['id_card']) < 15 or not rec['year_month']:
            continue

        # --- 脏数据检测：仅工伤>0且其他险种全部为0 → 跳过（非在职人员） ---
        other_total = (
            float(rec['pension_company_amount'] or 0)
            + float(rec['pension_employee_amount'] or 0)
            + float(rec['medical_company_amount'] or 0)
            + float(rec['medical_employee_amount'] or 0)
            + float(rec['maternity_company_amount'] or 0)
            + float(rec['unemployment_company_amount'] or 0)
            + float(rec['unemployment_employee_amount'] or 0)
            + float(rec['local_pension_company_amount'] or 0)
        )
        injury_only = float(rec['injury_company_amount'] or 0)
        if injury_only > 0 and other_total == 0:
            skipped_dirty.append(f'{rec["name"]}({rec["id_card"]}): 仅工伤¥{injury_only}')
            continue

        # 查找员工
        employee = None
        if rec['id_card']:
            employee = Employee.objects.filter(id_card=rec['id_card']).first()

        try:
            # 公积金：从员工档案读取，有则写入，无则为0
            housing_fund_employee = 0.0
            housing_fund_company = 0.0
            if employee and employee.has_housing_fund:
                # 优先级：新字段 housing_fund_employee > WageRecord > 旧字段 housing_fund_deduction
                if employee.housing_fund_employee and float(employee.housing_fund_employee) > 0:
                    housing_fund_employee = float(employee.housing_fund_employee)
                    housing_fund_company = float(employee.housing_fund_company or 0)
                else:
                    from apps.finance.models import WageRecord

                    wr = WageRecord.objects.filter(
                        employee=employee, year=rec['year_month'][:4], month=rec['year_month'][5:7]
                    ).first()
                    if wr and wr.housing_fund:
                        housing_fund_employee = float(wr.housing_fund)
                    elif employee.housing_fund_deduction:
                        housing_fund_employee = float(employee.housing_fund_deduction)
                    # 公司公积金 = 个人公积金 × 公司费率/个人费率（通常1:1）
                    if employee.company:
                        sc = getattr(employee.company, 'social_config', None)
                        if sc and sc.housing_fund_rate_employee and float(sc.housing_fund_rate_employee) > 0:
                            housing_fund_company = housing_fund_employee * (
                                float(sc.housing_fund_rate_company) / float(sc.housing_fund_rate_employee)
                            )
                        else:
                            housing_fund_company = housing_fund_employee  # 默认1:1

            float_vals = {
                'name': rec['name'],
                'pension_company': float(rec['pension_company_amount'] or 0),
                'pension_employee': float(rec['pension_employee_amount'] or 0),
                'medical_company': float(rec['medical_company_amount'] or 0),
                'medical_employee': float(rec['medical_employee_amount'] or 0),
                'pension_bup_company': float(rec['local_pension_company_amount'] or 0),
                'birth_company': float(rec['maternity_company_amount'] or 0),
                'injury_company': float(rec['injury_company_amount'] or 0),
                'unemployment_company': float(rec['unemployment_company_amount'] or 0),
                'unemployment_employee': float(rec['unemployment_employee_amount'] or 0),
                'housing_fund_employee': housing_fund_employee,
                'housing_fund_company': housing_fund_company,
                'is_reconciled': True,
            }
            # 去重查找：匹配唯一约束 (employee, year_month)
            # 如果未匹配到员工，则降级为 (company_id, id_card, year_month) 查找
            if employee:
                obj, is_new = SocialRecord.objects.update_or_create(
                    employee=employee,
                    year_month=rec['year_month'],
                    defaults={'company_id': actual_company_id, 'id_card': rec['id_card'], **float_vals},
                )
            else:
                obj, is_new = SocialRecord.objects.update_or_create(
                    company_id=actual_company_id,
                    id_card=rec['id_card'],
                    year_month=rec['year_month'],
                    defaults={**float_vals, 'employee': None, 'id_card': rec['id_card']},
                )
            if is_new:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f'员工 {rec["name"]}({rec["id_card"]}): {str(e)}')

    msg = f'导入完成，共处理 {len(records_to_save)} 条员工记录'
    if skipped_dirty:
        msg += f'，跳过 {len(skipped_dirty)} 条非在职人员（仅工伤）：{"; ".join(skipped_dirty)}'
    if errors:
        msg += f'，{len(errors)} 条记录有错误'

    return {
        'success': True,
        'message': msg,
        'created': created,
        'updated': updated,
        'skipped_dirty': skipped_dirty,
        'errors': errors[:50],  # 最多返回50条
    }


def _cell_raw(ws, row, col):
    """读取单元格值（col 为 0索引）"""
    try:
        return ws.cell(row=row, column=col + 1).value
    except (IndexError, TypeError):
        return None
