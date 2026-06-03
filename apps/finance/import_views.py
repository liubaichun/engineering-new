"""
Excel批量导入视图 - 财务模块
POST /api/finance/import/invoices/   - 发票导入
POST /api/finance/import/incomes/    - 收入导入
POST /api/finance/import/expenses/   - 支出导入
POST /api/finance/import/employees/   - 员工导入
"""

import datetime
import io
import re
from decimal import Decimal

from django.db import transaction
from openpyxl import load_workbook
from rest_framework.decorators import api_view
from apps.core.permissions import require_perms
from rest_framework.response import Response

from apps.finance.models import Invoice, Income, Expense, Employee, Company


def _parse_date(value) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(value) -> Decimal | None:
    if value is None or value == '':
        return None
    try:
        s = str(value).replace(',', '').replace('¥', '').replace(' ', '').strip()
        return Decimal(s)
    except (ValueError, TypeError):
        return None


def _parse_int(value) -> int | None:
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    v = str(value).strip().lower()
    return v in ('是', 'yes', 'true', '1', '有', '已', 'y', 't')


class ImportResult:
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors = []
        self.created_ids = []

    def add_error(self, row, field, msg, value=None):
        self.errors.append(
            {'row': row, 'field': field, 'message': msg, 'value': str(value) if value is not None else None}
        )
        self.error_count += 1

    def to_dict(self):
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'errors': self.errors,
            'created_ids': self.created_ids,
        }


def _get_col_map(ws, field_mapping: dict[str, str]) -> dict[str, int]:
    """根据表头行建立 field → col_index 映射"""
    headers = [c.value for c in ws[1]]
    col_map = {}
    for model_field, excel_header in field_mapping.items():
        if excel_header in headers:
            col_map[model_field] = headers.index(excel_header)
    return col_map


def _cell(ws, row_num: int, col_idx: int):
    if col_idx is None:
        return None
    try:
        return ws[row_num][col_idx].value
    except IndexError:
        return None


def _require(ws, row_num: int, col_idx: int, field_label: str, row_errors: list):
    val = _cell(ws, row_num, col_idx)
    if val is None or str(val).strip() == '':
        row_errors.append((field_label, f'必填字段 [{field_label}] 不能为空'))
    return val


# ─── 发票导入 ───────────────────────────────────────────────

INVOICE_FIELDS = {
    'invoice_no': '发票号码',
    'type': '发票类型',
    'invoice_type': '票种',
    'amount': '价税合计金额',
    'tax_rate': '税率(%)',
    'tax_amount': '税额',
    'counterparty': '对方单位',
    'is_credited': '是否认证',
    'status': '状态',
    'issue_date': '开票日期',
    'due_date': '到期日期',
    'remarks': '备注',
}


@api_view(['POST'])
@require_perms('finance:invoice:create')
def import_invoices(request):
    """
    导入发票。
    表头必须包含：发票号码、发票类型、价税合计金额、对方单位、开票日期
    """
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    col_map = _get_col_map(ws, INVOICE_FIELDS)
    missing = [v for f, v in INVOICE_FIELDS.items() if f not in col_map]
    if missing:
        return Response({'error': f'缺少必需列：{", ".join(missing)}'}, status=400)

    result = ImportResult()
    TYPE_MAP = {'收入发票': 'income', '支出发票': 'expense'}
    STATUS_MAP = {'待收款/待付款': 'pending', '已完成': 'paid', '已作废': 'cancelled'}

    for row_num in range(2, ws.max_row + 1):
        row_errors = []

        invoice_no = _cell(ws, row_num, col_map['invoice_no'])
        if not invoice_no:
            continue  # 跳过空行

        # 必填
        for f, label in [
            ('invoice_no', '发票号码'),
            ('type', '发票类型'),
            ('amount', '价税合计金额'),
            ('counterparty', '对方单位'),
            ('issue_date', '开票日期'),
        ]:
            v = _cell(ws, row_num, col_map[f])
            if not v:
                row_errors.append((label, f'必填字段 [{label}] 不能为空'))

        if row_errors:
            for field, msg in row_errors:
                result.add_error(row_num, field, msg)
            continue

        amount = _parse_decimal(_cell(ws, row_num, col_map['amount']))
        if amount is None:
            result.add_error(row_num, '价税合计金额', '金额格式错误')
            continue

        type_raw = str(_cell(ws, row_num, col_map['type'])).strip()
        type_val = TYPE_MAP.get(type_raw, 'income')

        tax_rate_raw = _cell(ws, row_num, col_map['tax_rate'])
        tax_rate = _parse_decimal(tax_rate_raw) or Decimal('0')
        tax_amount_raw = _cell(ws, row_num, col_map['tax_amount'])
        tax_amount = _parse_decimal(tax_amount_raw)

        issue_date = _parse_date(_cell(ws, row_num, col_map['issue_date']))
        due_date = _parse_date(_cell(ws, row_num, col_map['due_date']))

        status_raw = str(_cell(ws, row_num, col_map['status']) or 'pending').strip()
        status_val = STATUS_MAP.get(status_raw, 'pending')

        is_credited = _parse_bool(_cell(ws, row_num, col_map['is_credited']) or False)

        remarks = str(_cell(ws, row_num, col_map['remarks']) or '')

        # 自动计算不含税金额
        if tax_amount is None:
            if tax_rate and tax_rate > 0:
                tax_amount = (amount / (1 + tax_rate / 100) * tax_rate / 100).quantize(Decimal('0.01'))
            else:
                tax_amount = Decimal('0')

        try:
            with transaction.atomic():
                obj = Invoice.objects.create(
                    invoice_no=str(invoice_no).strip(),
                    type=type_val,
                    invoice_type=str(_cell(ws, row_num, col_map['invoice_type']) or '普通发票').strip(),
                    amount=amount,
                    tax_rate=tax_rate,
                    tax_amount=tax_amount,
                    counterparty=str(_cell(ws, row_num, col_map['counterparty'])).strip(),
                    is_credited=is_credited,
                    status=status_val,
                    issue_date=issue_date,
                    due_date=due_date,
                    remarks=remarks,
                    company=getattr(request, 'auth_company', None) or request.user.company,
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())


# ─── 收入导入 ───────────────────────────────────────────────

INCOME_FIELDS = {
    'amount': '金额',
    'date': '收入日期',
    'source': '来源',
    'status': '状态',
    'customer': '客户名称',
    'description': '描述/说明',
    'transaction_type': '交易类型',
    'summary': '摘要',
}


@api_view(['POST'])
@require_perms('finance:income:create')
def import_incomes(request):
    """导入收入记录。表头：金额、收入日期、客户名称"""
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    col_map = _get_col_map(ws, INCOME_FIELDS)
    missing = [v for f, v in INCOME_FIELDS.items() if f not in col_map]
    if missing:
        return Response({'error': f'缺少必需列：{", ".join(missing)}'}, status=400)

    result = ImportResult()
    STATUS_MAP = {'待审批': 'pending', '已批准': 'approved', '已拒绝': 'rejected'}

    for row_num in range(2, ws.max_row + 1):
        row_errors = []
        for f, label in [('amount', '金额'), ('date', '收入日期'), ('customer', '客户名称')]:
            v = _cell(ws, row_num, col_map[f])
            if not v:
                row_errors.append((label, '必填'))

        if row_errors:
            for field, msg in row_errors:
                result.add_error(row_num, field, msg)
            continue

        amount = _parse_decimal(_cell(ws, row_num, col_map['amount']))
        if amount is None:
            result.add_error(row_num, '金额', '金额格式错误')
            continue

        date = _parse_date(_cell(ws, row_num, col_map['date']))
        if date is None:
            result.add_error(row_num, '收入日期', '日期格式错误')
            continue

        customer = str(_cell(ws, row_num, col_map['customer']) or '').strip()
        source = str(_cell(ws, row_num, col_map['source']) or '').strip()
        status_val = STATUS_MAP.get(str(_cell(ws, row_num, col_map['status']) or '待审批').strip(), 'pending')
        description = str(_cell(ws, row_num, col_map['description']) or '').strip()
        transaction_type = str(_cell(ws, row_num, col_map.get('transaction_type')) or '').strip()
        summary = str(_cell(ws, row_num, col_map.get('summary')) or '').strip()

        try:
            with transaction.atomic():
                obj = Income.objects.create(
                    company=getattr(request, 'auth_company', None) or request.user.company,
                    amount=amount,
                    date=date,
                    source=source,
                    status=status_val,
                    customer=customer,
                    description=description,
                    transaction_type=transaction_type,
                    summary=summary,
                    operator=request.user,
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())


# ─── 支出导入 ───────────────────────────────────────────────

EXPENSE_FIELDS = {
    'amount': '金额',
    'date': '支出日期',
    'expense_type': '支出类型',
    'expense_category': '费用类别',
    'status': '状态',
    'supplier': '供应商/对方',
    'note': '摘要',
    'description': '详细说明',
    'transaction_type': '交易类型',
    'summary': '摘要2',
}


@api_view(['POST'])
@require_perms('finance:expense:create')
def import_expenses(request):
    """导入支出记录。表头：金额、支出日期、支出类型"""
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    col_map = _get_col_map(ws, EXPENSE_FIELDS)
    missing = [v for v in EXPENSE_FIELDS.values() if v not in [c.value for c in ws[1]]]
    if missing:
        return Response({'error': f'缺少必需列：{", ".join(missing)}'}, status=400)

    result = ImportResult()
    TYPE_MAP = {'费用报销': 'expense', '预付款': 'advance', '押金': 'deposit', '工资支出': 'wage'}
    STATUS_MAP = {'草稿': 'draft', '待审批': 'pending', '已批准': 'approved', '已拒绝': 'rejected'}

    for row_num in range(2, ws.max_row + 1):
        row_errors = []
        for f, label in [('amount', '金额'), ('expense_type', '支出类型')]:
            v = _cell(ws, row_num, col_map.get(f))
            if not v:
                row_errors.append((label, '必填'))

        if row_errors:
            for field, msg in row_errors:
                result.add_error(row_num, field, msg)
            continue

        amount = _parse_decimal(_cell(ws, row_num, col_map.get('amount')))
        if amount is None:
            result.add_error(row_num, '金额', '金额格式错误')
            continue

        type_raw = str(_cell(ws, row_num, col_map.get('expense_type') or '')).strip()
        type_val = TYPE_MAP.get(type_raw, 'expense')
        status_val = STATUS_MAP.get(str(_cell(ws, row_num, col_map.get('status')) or '待审批').strip(), 'pending')

        date = _parse_date(_cell(ws, row_num, col_map.get('date'))) or datetime.date.today()
        transaction_type = str(_cell(ws, row_num, col_map.get('transaction_type')) or '').strip()
        summary = str(_cell(ws, row_num, col_map.get('summary')) or '').strip()

        try:
            with transaction.atomic():
                obj = Expense.objects.create(
                    company=getattr(request, 'auth_company', None) or request.user.company,
                    amount=amount,
                    date=date,
                    expense_type=type_val,
                    expense_category=str(_cell(ws, row_num, col_map.get('expense_category')) or '').strip(),
                    status=status_val,
                    supplier=str(_cell(ws, row_num, col_map.get('supplier')) or '').strip(),
                    note=str(_cell(ws, row_num, col_map.get('note')) or '').strip(),
                    description=str(_cell(ws, row_num, col_map.get('description')) or '').strip(),
                    transaction_type=transaction_type,
                    summary=summary,
                    operator=request.user,
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())


# ─── 员工导入 ───────────────────────────────────────────────

EMPLOYEE_FIELDS = {
    'code': '工号',
    'name': '姓名',
    'id_card': '身份证号',
    'phone': '手机号码',
    'bank_card': '银行卡号',
    'bank_name': '开户行',
    'department': '部门',
    'position': '职位',
    'hire_date': '入职日期',
    'has_social_insurance': '是否购买社保',
    'has_housing_fund': '是否购买公积金',
    'social_insurance_base': '社保基数',
    'housing_fund_base': '公积金基数',
    'email': '邮箱',
    'emergency_contact': '紧急联系人',
    'emergency_phone': '紧急联系电话',
    'remarks': '备注',
}


@api_view(['POST'])
@require_perms('finance:employee:create')
def import_employees(request):
    """导入员工。表头：工号、姓名、身份证号、手机号码、入职日期"""
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    headers = [c.value for c in ws[1]]
    col_map = {}
    for model_field, excel_header in EMPLOYEE_FIELDS.items():
        if excel_header in headers:
            col_map[model_field] = headers.index(excel_header)

    # 找到公司名列（如果有）
    company_col = headers.index('所属公司') if '所属公司' in headers else None
    status_col = headers.index('状态') if '状态' in headers else None

    result = ImportResult()

    for row_num in range(2, ws.max_row + 1):
        row_errors = []
        for f, label in [('code', '工号'), ('name', '姓名')]:
            v = _cell(ws, row_num, col_map.get(f))
            if not v:
                row_errors.append((label, '必填'))

        if row_errors:
            for field, msg in row_errors:
                result.add_error(row_num, field, msg)
            continue

        name = str(_cell(ws, row_num, col_map['name'])).strip()
        code = str(_cell(ws, row_num, col_map['code'])).strip()
        id_card = str(_cell(ws, row_num, col_map.get('id_card')) or '').strip()
        phone = str(_cell(ws, row_num, col_map.get('phone')) or '').strip()
        # 清洗手机号
        phone = re.sub(r'\D', '', phone) if phone else ''

        hire_date = _parse_date(_cell(ws, row_num, col_map.get('hire_date')))
        if hire_date is None:
            result.add_error(row_num, '入职日期', '日期格式错误')
            continue

        # 公司
        company = getattr(request, 'auth_company', None) or request.user.company
        if company_col is not None:
            company_name = str(_cell(ws, row_num, company_col) or '').strip()
            if company_name:
                company = Company.objects.filter(name=company_name).first()

        # 状态
        emp_status = 'active'
        if status_col is not None:
            s = str(_cell(ws, row_num, status_col) or '在职').strip()
            if s in ('离职', '待入职', '实习'):
                emp_status = 'inactive' if s == '离职' else 'probation'

        # 社保/公积金
        has_si = _parse_bool(_cell(ws, row_num, col_map.get('has_social_insurance') or False))
        has_hf = _parse_bool(_cell(ws, row_num, col_map.get('has_housing_fund') or False))
        si_base = _parse_decimal(_cell(ws, row_num, col_map.get('social_insurance_base')))
        hf_base = _parse_decimal(_cell(ws, row_num, col_map.get('housing_fund_base')))

        try:
            with transaction.atomic():
                obj = Employee.objects.create(
                    code=code,
                    name=name,
                    id_card=id_card,
                    phone=phone,
                    bank_card=str(_cell(ws, row_num, col_map.get('bank_card')) or '').strip(),
                    bank_name=str(_cell(ws, row_num, col_map.get('bank_name')) or '').strip(),
                    department=str(_cell(ws, row_num, col_map.get('department')) or '').strip(),
                    position=str(_cell(ws, row_num, col_map.get('position')) or '').strip(),
                    company=company,
                    hire_date=hire_date,
                    status=emp_status,
                    has_social_insurance=has_si,
                    has_housing_fund=has_hf,
                    social_insurance_base=si_base or Decimal('0'),
                    housing_fund_base=hf_base or Decimal('0'),
                    email=str(_cell(ws, row_num, col_map.get('email')) or '').strip(),
                    emergency_contact=str(_cell(ws, row_num, col_map.get('emergency_contact')) or '').strip(),
                    emergency_phone=re.sub(
                        r'\D', '', str(_cell(ws, row_num, col_map.get('emergency_phone')) or '').strip()
                    ),
                    remarks=str(_cell(ws, row_num, col_map.get('remarks')) or '').strip(),
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())
