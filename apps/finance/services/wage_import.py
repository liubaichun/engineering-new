"""
工资批量导入服务
支持 Excel（.xlsx）格式，一行一条工资记录
"""

from __future__ import annotations

import openpyxl
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple
from io import BytesIO


class WageImportError(Exception):
    """导入异常，可携带行号信息"""

    def __init__(self, message: str, row: Optional[int] = None, field: Optional[str] = None) -> None:
        super().__init__(message)
        self.row = row
        self.field = field


# 列名 → 字段名 映射（不区分大小写，支持多种写法）
COLUMN_MAP = {
    # 员工识别
    '姓名': 'employee_name',
    '员工姓名': 'employee_name',
    '姓名/工号': 'employee_name',
    # 期间
    '年份': 'year',
    '年度': 'year',
    '月份': 'month',
    '月': 'month',
    # 应发
    '基本工资': 'base_salary',
    '岗位工资': 'position_salary',
    '加班费': 'overtime_pay',
    '奖金': 'bonus',
    '提成': 'commission',
    '餐补': 'meal_allowance',
    '交通补贴': 'transport_allowance',
    '通讯补贴': 'communication_allowance',
    '其他应发': 'other_allowance',
    # 社保公积金（个人部分）
    '社保扣款': 'social_insurance',
    '公积金': 'housing_fund',
    '公积金扣款': 'housing_fund',
    # 其他扣款
    '请假天数': 'leave_days',
    '请假扣款': 'leave_deduction',
    '病假天数': 'sick_leave_days',
    '病假扣款': 'sick_leave_deduction',
    '员工借款还款': 'employee_loan_repayment',
    '其他借款': 'other_loan',
    '其他扣款': 'other_deductions',
    # 银行卡
    '银行卡号': 'bank_card',
    '卡号': 'bank_card',
}


def parse_decimal(value: Any) -> Optional[Decimal]:
    """将单元格值转为 Decimal，失败返回 None"""
    if value is None or value == '':
        return Decimal('0')
    try:
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        s = str(value).strip().replace(',', '').replace('¥', '').replace('元', '')
        if s == '' or s == '-':
            return Decimal('0')
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_int(value: Any) -> Optional[int]:
    """将单元格值转为 int，失败返回 None"""
    if value is None or value == '':
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, InvalidOperation):
        return None


def _normalize_header(cell_value: str) -> str:
    """标准化列头文字"""
    if cell_value is None:
        return ''
    return str(cell_value).strip()


def import_wage_excel(file_bytes: bytes, company_id: int, defaults: Optional[Dict[str, Any]] = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    解析工资导入 Excel，返回 (成功记录列表, 错误列表)

    返回格式：
    - 成功: [{employee_name, year, month, base_salary, ...}, ...]
    - 错误: [{row: int, message: str, field: str}, ...]

    参数:
        file_bytes: Excel 文件二进制内容
        company_id: 导入到的公司 ID
        defaults: 默认值 dict（如 {'year': 2026, 'month': 5}）
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 第一行是表头
    headers = []
    for cell in ws[1]:
        headers.append(_normalize_header(cell.value))

    # 建立列索引
    col_index = {}  # field_name -> col_index
    for idx, h in enumerate(headers):
        if h in COLUMN_MAP:
            field = COLUMN_MAP[h]
            if field not in col_index:  # 避免重复列覆盖
                col_index[field] = idx

    # 必须有的列
    required = ['employee_name']
    missing = [f for f in required if f not in col_index]
    if missing:
        raise WageImportError(f'缺少必需列：{", ".join(missing)}')

    records = []
    errors = []

    for row_num in range(2, ws.max_row + 1):
        # 跳过空行
        row_values = [ws.cell(row=row_num, column=idx + 1).value for idx in range(len(headers))]
        if all(v is None or str(v).strip() == '' for v in row_values):
            continue

        try:
            record = {'company_id': company_id}
            if defaults:
                record.update(defaults)

            for field, col_idx in col_index.items():
                raw_val = row_values[col_idx] if col_idx < len(row_values) else None
                if field in ('year', 'month'):
                    val = parse_int(raw_val)
                    if val is None:
                        raise WageImportError(f"'{field}' 值无效：{raw_val}", row=row_num, field=field)
                elif field == 'employee_name':
                    val = str(raw_val).strip() if raw_val else ''
                    if not val:
                        raise WageImportError('员工姓名为空', row=row_num, field='employee_name')
                elif field in ('leave_days', 'sick_leave_days'):
                    val = parse_int(raw_val)
                    record[field] = val if val is not None else 0
                    continue
                else:
                    val = parse_decimal(raw_val)
                    if val is None and field not in ('bank_card', 'employee_name'):
                        raise WageImportError(f"'{field}' 值无效：{raw_val}", row=row_num, field=field)
                    val = val or Decimal('0')

                record[field] = val

            records.append(record)

        except WageImportError:
            raise
        except Exception as e:
            errors.append({'row': row_num, 'message': str(e), 'field': ''})

    return records, errors


def preview_wage_excel(file_bytes: bytes, company_id: int) -> Dict[str, Any]:
    """
    预览 Excel 内容，不入库。返回前 10 行数据 + 字段列表 + 错误提示。
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [_normalize_header(c.value) for c in ws[1]]
    preview_rows = []
    for row_num in range(2, min(ws.max_row + 1, 12)):
        row_values = [ws.cell(row=row_num, column=idx + 1).value for idx in range(len(headers))]
        if all(v is None or str(v).strip() == '' for v in row_values):
            continue
        preview_rows.append(row_values)

    # 字段映射预览
    field_preview = {}
    for idx, h in enumerate(headers):
        if h in COLUMN_MAP:
            field_preview[h] = COLUMN_MAP[h]

    return {
        'headers': headers,
        'field_map': field_preview,
        'preview_rows': preview_rows,
        'total_rows': ws.max_row - 1,
    }
