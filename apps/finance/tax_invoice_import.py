"""
税务局电子发票批量导入模块
支持：进项发票（取得发票）→ 生成 Expense
      销项发票（开出发票）→ 生成 Income

税务局导出格式（信息汇总表 Sheet）：
  序号 | 发票代码 | 发票号码 | 数电发票号码 | 销方识别号 | 销方名称 | 购方识别号 | 购买方名称
  | 开票日期 | 税收分类编码 | 特定业务类型 | 货物或应税劳务名称 | 规格型号 | 单位 | 数量 | 单价
  | 金额 | 税率 | 税额 | 价税合计 | 发票来源 | 发票票种 | 发票状态 | 是否正数发票 | 发票风险等级 | 开票人 | 备注
"""
import datetime
import io
import re
from decimal import Decimal

from django.db import transaction
from openpyxl import load_workbook
from rest_framework import permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from apps.crm.models import Client, Supplier
from apps.finance.models import Company, Expense, Income


# ─── 工具函数 ───────────────────────────────────────────────

def _parse_date(value) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(value) -> Decimal | None:
    if value is None or value == '':
        return None
    try:
        s = str(value).replace(',', '').replace('¥', '').replace(' ', '').strip()
        if s == '' or s == '-':
            return None
        return Decimal(s)
    except (ValueError, TypeError):
        return None


def _cell(ws, row_num: int, col_idx: int):
    """安全读取单元格，col_idx 从 0 开始"""
    if col_idx is None:
        return None
    try:
        return ws[row_num][col_idx].value
    except (IndexError, TypeError):
        return None


def _is_enterprise_name(name: str) -> bool:
    """
    判断对方名称是否为企业（有正规税号/公司名）。
    过滤：税务局、国库、公积金管理中心、无名称等非企业。
    """
    if not name:
        return False
    name = name.strip()
    # 非企业关键字
    skip_patterns = [
        r'国家金库', r'税务局', r'国家税务总局', r'住房公积金',
        r'中华人民共和国', r'深圳市.*金库', r'对公中间业务',
        r'暂收款', r'应付利息', r'应收利息', r'应付账款',
        r'自动驾驶', r'测试', r'示范应用', r'备付金',
    ]
    for p in skip_patterns:
        if re.search(p, name):
            return False
    # 企业名通常含公司/有限/集团/科技/贸易等
    enterprise_patterns = [
        r'公司', r'有限公司', r'有限责任公司', r'集团', r'科技', r'贸易',
        r'实业', r'发展', r'公司$', r'Co', r'Ltd', r'Inc',
    ]
    for p in enterprise_patterns:
        if re.search(p, name):
            return True
    return False


def _is_government_or_special(name: str) -> bool:
    """判断是否为政府/事业单位"""
    if not name:
        return False
    patterns = [r'大学', r'医院', r'政府', r'委员会', r'局$', r'厅$', r'处$',
                r'事业单位', r'法院', r'检察院', r'管委会', r'政府$']
    for p in patterns:
        if re.search(p, name):
            return True
    return False


def _classify_counterparty(name: str) -> str:
    """分类：enterprise / government / individual"""
    if _is_government_or_special(name):
        return 'government'
    if _is_enterprise_name(name):
        return 'enterprise'
    return 'individual'


def _classify_expense_category(summary: str, goods_name: str) -> str:
    """
    根据摘要/商品名称自动分类 expense_category
    """
    text = f"{summary} {goods_name}".lower()
    if '工资' in text or '代发' in text or '奖金' in text:
        return '工资'
    if '税' in text or '增值税' in text or '个税' in text or '所得税' in text or '税费' in text:
        return '税费'
    if '社保' in text or '公积金' in text or '保险' in text:
        return '社保公积金'
    if '交通' in text or '油' in text or '过路' in text or '停车' in text or '打车' in text or '滴滴' in text:
        return '交通差旅'
    if '餐' in text or '吃饭' in text or '招待' in text or '宴请' in text:
        return '业务招待'
    if '京东' in text or '商城' in text or '天猫' in text or '淘宝' in text or '电商' in text:
        return '办公用品/网上采购'
    if '电脑' in text or '软件' in text or '服务器' in text or '网络' in text or '云' in text:
        return '技术服务'
    if '办公' in text or '文具' in text or '耗材' in text or '纸张' in text:
        return '办公费用'
    if '维修' in text or '维护' in text or '保养' in text or '修理' in text:
        return '维修维护'
    if '水' in text or '电' in text or '煤气' in text or '燃气' in text or '物业' in text:
        return '水电物业'
    if '广告' in text or '宣传' in text or '推广' in text or '营销' in text:
        return '营销推广'
    if '咨询' in text or '顾问' in text or '审计' in text or '法律' in text:
        return '咨询顾问'
    if '银行' in text or '手续费' in text or '结算' in text:
        return '金融服务'
    if '租赁' in text or '租金' in text or '房租' in text:
        return '租赁费'
    if '快递' in text or '运输' in text or '物流' in text or '仓储' in text:
        return '物流快递'
    return '其他费用'


def _classify_income_category(summary: str, goods_name: str) -> str:
    """根据摘要/商品名称自动分类收入类别"""
    text = f"{summary} {goods_name}".lower()
    if '软件' in text or '维护' in text or '技术服务' in text or '开发' in text:
        return '技术服务费'
    if '咨询' in text or '顾问' in text:
        return '咨询服务费'
    if '培训' in text or '教育' in text:
        return '培训费'
    if '设备' in text or '硬件' in text or '器材' in text:
        return '设备销售'
    if '租赁' in text or '租金' in text:
        return '租赁收入'
    if '退款' in text:
        return '退款'
    if '利息' in text:
        return '利息收入'
    return '其他收入'


# ─── 核心导入函数 ───────────────────────────────────────────

class TaxInvoiceImportResult:
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.skipped_count = 0
        self.errors = []
        self.created_expense_ids = []
        self.created_income_ids = []
        self.skipped = []  # [(row, reason)]

    def add_error(self, row, field, msg, value=None):
        self.errors.append({'row': row, 'field': field, 'message': msg,
                            'value': str(value)[:50] if value else None})
        self.error_count += 1

    def add_skipped(self, row, reason):
        self.skipped.append({'row': row, 'reason': reason})
        self.skipped_count += 1

    def to_dict(self):
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'skipped_count': self.skipped_count,
            'errors': self.errors[:50],  # 最多50条
            'skipped': self.skipped[:100],
            'created_expense_ids': self.created_expense_ids,
            'created_income_ids': self.created_income_ids,
        }


def _import_tax_invoices_common(ws, company: Company, invoice_type: str, operator):
    """
    通用发票导入逻辑。

    invoice_type: 'expense' = 进项票（供应商开给我们的）→ Expense
                  'income'  = 销项票（我们开给客户的）→ Income
    """
    result = TaxInvoiceImportResult()

    # 读取表头（Row 1）
    headers = [c.value for c in ws[1]]
    if not headers or headers[0] is None:
        result.add_error(1, '表头', '文件格式错误，无法读取表头')
        return result

    def get_col(name: str) -> int | None:
        """找列索引，支持全角半角变体"""
        for h in headers:
            if h and name in str(h):
                return headers.index(h)
        return None

    # 建立列索引映射
    idx = {
        'invoice_no': get_col('发票号码'),
        'elec_invoice_no': get_col('数电发票号码'),
        'seller_tax_id': get_col('销方识别号'),
        'seller_name': get_col('销方名称'),
        'buyer_tax_id': get_col('购方识别号'),
        'buyer_name': get_col('购买方名称'),
        'issue_date': get_col('开票日期'),
        'goods_name': get_col('货物或应税劳务名称'),
        'amount': get_col('金额'),        # 不含税金额
        'tax_rate': get_col('税率'),
        'tax_amount': get_col('税额'),
        'total_amount': get_col('价税合计'),
        'invoice_type_name': get_col('发票票种'),
        'status_name': get_col('发票状态'),
        'is_positive': get_col('是否正数发票'),
        'invoice_source': get_col('发票来源'),
        'issuer': get_col('开票人'),
        'remarks': get_col('备注'),
    }

    # 验证必需列
    missing = [k for k, v in idx.items() if v is None and k not in ('elec_invoice_no', 'remarks', 'issuer', 'invoice_source', 'invoice_type_name')]
    if missing:
        for k in missing:
            result.add_error(1, k, f'缺少列：{k}')
        return result

    # 按发票号分组（一张发票可能有多行商品）
    invoice_groups = {}  # invoice_no -> {meta, rows}
    for row_num in range(2, len(ws) + 1):
        # 取发票号（优先数电发票号码，其次普通发票号码）
        elec_no = str(_cell(ws, row_num, idx['elec_invoice_no']) or '').strip()
        normal_no = str(_cell(ws, row_num, idx['invoice_no']) or '').strip()
        invoice_no = elec_no or normal_no
        if not invoice_no or invoice_no == 'None':
            continue

        # 过滤非正数发票
        is_positive = str(_cell(ws, row_num, idx['is_positive']) or '是').strip()
        if is_positive not in ('是', 'yes', 'true', '1'):
            result.add_skipped(row_num, f'负数发票，跳过')
            continue

        # 过滤非正常发票
        status = str(_cell(ws, row_num, idx['status_name']) or '正常').strip()
        if status != '正常':
            result.add_skipped(row_num, f'发票状态[{status}]，跳过')
            continue

        if invoice_no not in invoice_groups:
            invoice_groups[invoice_no] = {
                'meta': {k: _cell(ws, row_num, v) for k, v in idx.items() if v is not None},
                'rows': []
            }
        invoice_groups[invoice_no]['rows'].append(row_num)

    # 逐张处理
    for invoice_no, group in invoice_groups.items():
        row = group['rows'][0]
        meta = group['meta']

        # 解析金额
        total_amount = _parse_decimal(meta.get('total_amount'))
        tax_amount = _parse_decimal(meta.get('tax_amount'))
        amount = _parse_decimal(meta.get('amount'))
        # 如果价税合计有值，用它作为实际发生额
        if total_amount is None or total_amount == 0:
            result.add_skipped(row, f'价税合计为空，跳过')
            continue
        actual_amount = total_amount  # 实际付款/收款金额

        # 解析税率
        tax_rate_raw = meta.get('tax_rate')
        tax_rate = _parse_decimal(tax_rate_raw) or Decimal('0')
        # 税率可能是 "13%" 字符串
        if isinstance(tax_rate_raw, str) and '%' in tax_rate_raw:
            try:
                tax_rate = Decimal(tax_rate_raw.replace('%', '').strip())
            except (ValueError, Decimal.InvalidOperation):
                tax_rate = Decimal('0')

        # 解析日期
        issue_date = _parse_date(meta.get('issue_date'))
        if issue_date is None:
            issue_date = datetime.date.today()

        # 商品名称（合并多行）
        goods_list = []
        for r in group['rows']:
            g = str(_cell(ws, r, idx['goods_name']) or '').strip()
            if g:
                goods_list.append(g)
        goods_name = '；'.join(goods_list[:3])  # 最多3个商品

        # 摘要/备注
        remarks = str(meta.get('remarks') or '').strip()
        summary = goods_name[:100] if goods_name else ''

        # 来源票种
        invoice_type_name = str(meta.get('invoice_type_name') or '数电发票（普通发票）').strip()
        invoice_source = str(meta.get('invoice_source') or '').strip()
        issuer = str(meta.get('issuer') or '').strip()

        # 验证：进项票必须有销方名称，销项票必须有购方名称（在 atomic 外检查，避免事务污染）
        seller_name = str(meta.get('seller_name') or '').strip()
        buyer_name = str(meta.get('buyer_name') or '').strip()
        if invoice_type == 'expense' and not seller_name:
            result.add_skipped(row, '销方名称为空，跳过')
            continue
        if invoice_type != 'expense' and not buyer_name:
            result.add_skipped(row, '购买方名称为空，跳过')
            continue

        try:
            with transaction.atomic():
                if invoice_type == 'expense':
                    # 进项票：供应商开给我们 → 支出
                    seller_tax_id = str(meta.get('seller_tax_id') or '').strip()
                    buyer_tax_id = str(meta.get('buyer_tax_id') or '').strip()  # 我们公司

                    cp_type = _classify_counterparty(seller_name)
                    category = _classify_expense_category(summary, goods_name)

                    # 自动创建/更新 Supplier
                    if cp_type == 'enterprise':
                        supplier_obj, created = Supplier.objects.get_or_create(
                            company=company,
                            name=seller_name,
                            defaults={
                                'counterparty_type': cp_type,
                                'tax_id': seller_tax_id,
                                'created_by': operator,
                            }
                        )
                        if not supplier_obj.tax_id and seller_tax_id:
                            supplier_obj.tax_id = seller_tax_id
                            supplier_obj.save(update_fields=['tax_id'])
                        if not supplier_obj.bank_account:
                            # 从备注里尝试提取银行账号
                            bank_match = re.search(r'账号[：:]\s*(\d{10,})', remarks)
                            if bank_match:
                                supplier_obj.bank_account = bank_match.group(1)
                                supplier_obj.save(update_fields=['bank_account'])

                    counterparty_display = seller_name

                    obj = Expense.objects.create(
                        company=company,
                        amount=actual_amount,
                        date=issue_date,
                        expense_date=issue_date,
                        expense_type='expense',
                        expense_category=category,
                        supplier=seller_name,
                        note=f"[进项票]{invoice_no} {invoice_type_name}",
                        description=f"{goods_name}\n票号:{invoice_no}\n开票人:{issuer}\n来源:{invoice_source}".strip(),
                        operator=operator,
                        status='approved',
                        transaction_type=f"进项票-{invoice_type_name}",
                        summary=summary,
                    )
                    result.created_expense_ids.append(obj.id)

                else:
                    # 销项票：我们开给客户 → 收入
                    buyer_tax_id = str(meta.get('buyer_tax_id') or '').strip()
                    seller_tax_id = str(meta.get('seller_tax_id') or '').strip()  # 我们公司

                    cp_type = _classify_counterparty(buyer_name)
                    category = _classify_income_category(summary, goods_name)

                    # 自动创建/更新 Client
                    if cp_type in ('enterprise', 'government'):
                        client_obj, created = Client.objects.get_or_create(
                            company=company,
                            name=buyer_name,
                            defaults={
                                'counterparty_type': cp_type,
                                'tax_id': buyer_tax_id,
                                'created_by': operator,
                            }
                        )
                        if not client_obj.tax_id and buyer_tax_id:
                            client_obj.tax_id = buyer_tax_id
                            client_obj.save(update_fields=['tax_id'])

                    counterparty_display = buyer_name

                    obj = Income.objects.create(
                        company=company,
                        amount=actual_amount,
                        date=issue_date,
                        source=f'发票导入[{invoice_no}]',
                        customer=buyer_name,
                        description=f"{goods_name}\n票号:{invoice_no}\n开票人:{issuer}\n税率:{tax_rate}%".strip(),
                        operator=operator,
                        status='approved',
                        transaction_type=f"销项票-{invoice_type_name}",
                        summary=summary,
                    )
                    result.created_income_ids.append(obj.id)

                result.success_count += 1

        except Exception as ex:
            result.add_error(row, 'general', f'保存失败: {str(ex)}')

    return result


# ─── API 视图 ───────────────────────────────────────────────


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_tax_invoices(request):
    """
    导入税务发票（进项/销项通用）

    POST 参数：
      file: .xlsx 文件（税务局导出格式，信息汇总表 Sheet）
      invoice_type: 'expense'（进项票） 或 'income'（销项票）
      company_id: 公司ID
    """
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    invoice_type = request.POST.get('invoice_type', 'expense')
    if invoice_type not in ('expense', 'income'):
        return Response({'error': "invoice_type 必须是 'expense'（进项）或 'income'（销项）"}, status=400)

    company_id = request.POST.get('company_id')
    if not company_id:
        return Response({'error': '缺少 company_id 参数'}, status=400)

    try:
        company = Company.objects.get(id=company_id)
    except Company.DoesNotExist:
        return Response({'error': f'公司ID {company_id} 不存在'}, status=404)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active

    # 找"信息汇总表" Sheet
    sheet_name = ws.title
    if sheet_name != '信息汇总表':
        # 尝试找对应的 Sheet
        for s in wb.sheetnames:
            if '信息汇总' in s or '汇总' in s:
                ws = wb[s]
                break

    result = _import_tax_invoices_common(ws, company, invoice_type, request.user)
    return Response(result.to_dict())


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_tax_invoices_auto(request):
    """
    自动识别并导入进项+销项发票（一次性处理整个文件）

    POST 参数：
      file: .xlsx 文件（税务局导出格式）
      company_id: 公司ID

    说明：进项票和销项票通常是两个不同的文件，此接口用于
          兼容两个文件分别上传的场景。
    """
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    company_id = request.POST.get('company_id')
    if not company_id:
        return Response({'error': '缺少 company_id 参数'}, status=400)

    try:
        company = Company.objects.get(id=company_id)
    except Company.DoesNotExist:
        return Response({'error': f'公司ID {company_id} 不存在'}, status=404)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)

    # 按 Sheet 分离进项/销项
    # 通常进项票 Sheet 名含"取得"，销项票含"开具"
    # 但如果只有一个 Sheet，根据内容判断
    results = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        # 取第一行检查购方/销方
        first_row = [c.value for c in ws[1]] or []
        first_str = ' '.join([str(v) for v in first_row if v])

        if '购买方' in first_str or '购方' in first_str or '销方' not in first_str:
            # 可能是进项票（购买方=我们公司）
            # 判断：我们公司的税号在购方列
            inv_type = 'expense'
        elif '销方' in first_str:
            # 可能是销项票（销方=我们公司）
            inv_type = 'income'
        else:
            continue  # 跳过不认识的 Sheet

        result = _import_tax_invoices_common(ws, company, inv_type, request.user)
        key = f'{inv_type}_{sname}'
        results[key] = result.to_dict()

    if not results:
        return Response({'error': '未识别到有效发票数据，请确认文件格式为税务局导出格式'}, status=400)

    total_success = sum(r['success_count'] for r in results.values())
    total_errors = sum(r['error_count'] for r in results.values())
    total_skipped = sum(r['skipped_count'] for r in results.values())

    return Response({
        'total_success': total_success,
        'total_errors': total_errors,
        'total_skipped': total_skipped,
        'by_sheet': results,
    })
