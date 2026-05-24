"""
银行流水解析适配器
支持：ICBC(工商银行) / CMB(招商银行) / CCB(建设银行) / BOC(中国银行) / ABC(农业银行) / COMM(交通银行) / PSBC(邮储银行)
"""
import datetime
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Optional


@dataclass
class ParsedTransaction:
    """只保留银行流水中实际有数据的12个字段"""
    transaction_date: Optional[datetime.date] = None
    transaction_time: Optional[datetime.time] = None
    amount: Decimal = Decimal('0')
    direction: str = 'expense'   # income / expense（贷=income，借=expense）
    balance: Optional[Decimal] = None
    bank_serial: str = ''       # 流水号
    counterparty_name: str = ''  # 收(付)方名称
    counterparty_account: str = ''  # 收(付)方账号
    counterparty_bank: str = ''  # 收(付)方开户行名
    summary: str = ''           # 摘要
    transaction_type: str = ''  # 交易类型（原始文本，不分类）
    account_no: str = ''        # 我方账号（从文件"账号"列提取，用于与账户归属校验）

    def get_dedup_key(self) -> str:
        if self.bank_serial:
            return self.bank_serial
        date_str = self.transaction_date.isoformat() if self.transaction_date else ''
        time_str = self.transaction_time.isoformat() if self.transaction_time else ''
        return f"{date_str}_{time_str}_{self.counterparty_account}_{self.amount}"


class BankStatementAdapter(ABC):
    bank_code: str = ''
    bank_name: str = ''

    @abstractmethod
    def detect(self, ws) -> bool:
        raise NotImplementedError

    @abstractmethod
    def parse(self, ws) -> list[ParsedTransaction]:
        raise NotImplementedError

    # ── 通用解析工具 ──────────────────────────────────────────────
    def _parse_date(self, value) -> Optional[datetime.date]:
        if value is None:
            return None
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        s = str(value).strip().replace('/', '-').replace('年', '-').replace('月', '-').replace('日', '')
        # 去掉时间部分，只保留日期
        s = s.split(' ')[0].split('T')[0]
        for fmt in ('%Y-%m-%d', '%Y%m%d', '%m-%d-%Y', '%d-%m-%Y', '%Y年%m月%d日', '%Y.%m.%d'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_time(self, value) -> Optional[datetime.time]:
        if value is None:
            return None
        if isinstance(value, datetime.time):
            return value
        if isinstance(value, datetime.datetime):
            return value.time()
        s = str(value).strip()
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return datetime.datetime.strptime(s, fmt).time()
            except ValueError:
                continue
        return None

    def _decimal(self, value) -> Optional[Decimal]:
        if value is None or str(value).strip() == '':
            return None
        try:
            s = str(value).replace(',', '').replace('¥', '').replace(' ', '').replace('+', '').replace('（', '').replace('）', '').strip()
            if s == '' or s == '-':
                return None
            return Decimal(s)
        except (ValueError, TypeError):
            return None


# ─── 工商银行 HISTORYDETAIL ─────────────────────────────────────────────
class ICBCAdapter(BankStatementAdapter):
    """
    格式：
    Row 1 (A1): [HISTORYDETAIL]
    Row 2: 表头行
    Row 3+: 数据行
    表头列名：凭证号、对方账号、交易时间、借贷标志、对方单位、对方行号、用途、摘要、附言、回单个性化信息、余额、时间戳、发生额、入账日期、入账时间、本方账号、转出金额、转入金额
    借贷标志：贷=收入（从"转入金额"取金额），借=支出（从"转出金额"取金额）
    余额列存在时直接使用，不依赖余额变化推导金额
    """
    bank_code = 'ICBC'
    bank_name = '工商银行'

    def detect(self, ws) -> bool:
        try:
            return str(ws.cell(1, 1).value or '').strip() == '[HISTORYDETAIL]'
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        # 动态查找表头行
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if any('借贷' in v or '借/贷' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            return []

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    val = ws.cell(row, c).value
                    if val is not None:
                        return val
                # 模糊匹配
                for h, ci in col_idx.items():
                    if name in h:
                        v = ws.cell(row, ci).value
                        if v is not None:
                            return v
            return None

        # 从第一行数据提取本方账号（"本方账号"列），供所有记录使用
        file_account_no = ''
        first_data_row = header_row + 1
        if first_data_row <= ws.max_row:
            raw_acct = get_col(first_data_row, '本方账号', '本公司账号', '我的账号')
            if raw_acct:
                file_account_no = str(raw_acct).strip()

        # 直接读文件已有字段，1:1记录，不推算
        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                direction_flag = str(get_col(row_idx, '借贷标志', '借贷', '借/贷') or '').strip()
                # 借贷标志直接决定方向，不推算
                if direction_flag == '贷':
                    direction = 'income'
                elif direction_flag == '借':
                    direction = 'expense'
                else:
                    continue  # 无借贷标志的行跳过

                txn_date = self._parse_date(
                    get_col(row_idx, '入账日期', '交易日期', '记账日期', '日期') or
                    get_col(row_idx, '入账时间')
                )
                if not txn_date:
                    continue

                # 入账时间 → transaction_time（优先用入账时间，不用交易时间）
                txn_time = self._parse_time(get_col(row_idx, '入账时间')) or \
                           self._parse_time(get_col(row_idx, '交易时间'))

                # 金额：根据借贷方向从不同列读取
                # 借方（支出）从"转出金额"取，贷方（收入）从"转入金额"取
                if direction == 'expense':
                    amount_raw = get_col(row_idx, '转出金额')
                    amount = self._decimal(amount_raw) if amount_raw else None
                else:
                    amount_raw = get_col(row_idx, '转入金额')
                    amount = self._decimal(amount_raw) if amount_raw else None

                # 余额：直接读余额列
                balance_raw = get_col(row_idx, '余额', '账户余额', '可用余额')
                balance = self._decimal(balance_raw) if balance_raw else None

                # 用途 → 交易类型
                transaction_type = str(get_col(row_idx, '用途') or '').strip()

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=txn_time,
                    amount=amount,
                    direction=direction,
                    balance=balance,
                    counterparty_name=str(get_col(row_idx, '对方单位', '对方名称', '对方户名', '收(付)方名称') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '对方账号', '对方账户', '账号') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '对方行号', '对方银行', '开户行') or '').strip(),
                    summary=str(get_col(row_idx, '摘要', '交易描述', '说明') or '').strip(),
                    bank_serial=str(get_col(row_idx, '凭证号', '流水号', '交易流水') or '').strip(),
                    transaction_type=transaction_type,
                    account_no=file_account_no,
                ))
            except Exception:
                continue

        return records

# ─── 招商银行 ACCOUNT v2.0 ──────────────────────────────────────────────
class CMBAdapter(BankStatementAdapter):
    """
    格式（v2.0，对账单格式）：
    前若干行为元数据，第N行(索引N)为表头（含'交易日'+'流水号'），第N+1行开始为数据
    表头：交易日|交易时间|流水号|借方金额|贷方金额|余额|收(付)方名称|收(付)方账号|收(付)方开户行名|摘要|其它摘要
    """
    bank_code = 'CMB'
    bank_name = '招商银行'

    def detect(self, ws) -> bool:
        try:
            # CMB v2.0 特征：Row5='对账单'(col2), Row6='接口版本' 2.0
            for row in range(1, 15):
                for col in range(1, 10):
                    val = str(ws.cell(row, col).value or '')
                    if val == '对账单':
                        for c in range(1, 6):
                            v2 = str(ws.cell(row + 1, c).value or '')
                            if v2 == '2.0':
                                return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        """
        只解析银行流水文件中实际有数据的11个字段：
        交易日、交易时间、借方金额、贷方金额、余额、摘要、流水号、
        收(付)方名称、收(付)方账号、收(付)方开户行名、交易类型
        """
        # 找表头行（含'交易日'+'流水号'）
        header_row = None
        for r in range(1, ws.max_row + 1):
            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1)]
            if any('交易日' in v for v in row_vals) and any('流水号' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            return []

        # 动态匹配列名（兼容全角/半角括号）
        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, name):
            for variant in [name, name.replace('(', '（').replace(')', '）'),
                            name.replace('（', '(').replace('）', ')')]:
                c = col_idx.get(variant)
                if c:
                    val = ws.cell(row, c).value
                    if val is not None:
                        return val
            return None

        # 提前提取文件中的账号（第1列"账号"），用于与账户归属校验
        file_account_no = ''
        for r in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            acc = str(get_col(r, '账号') or '').strip()
            if acc:
                file_account_no = acc
                break

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                debit  = self._decimal(get_col(row_idx, '借方金额'))
                credit = self._decimal(get_col(row_idx, '贷方金额'))

                if debit is not None and debit > 0:
                    amount, direction = debit, 'expense'
                elif credit is not None and credit > 0:
                    amount, direction = credit, 'income'
                else:
                    continue

                txn_date = self._parse_date(get_col(row_idx, '交易日'))
                if not txn_date:
                    continue

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=self._parse_time(get_col(row_idx, '交易时间')),
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '余额')),
                    bank_serial=str(get_col(row_idx, '流水号') or '').strip(),
                    counterparty_name=str(get_col(row_idx, '收(付)方名称') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '收(付)方账号') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '收(付)方开户行名') or '').strip(),
                    summary=str(get_col(row_idx, '摘要') or '').strip(),
                    transaction_type=str(get_col(row_idx, '交易类型') or '').strip(),
                    account_no=file_account_no,
                ))
            except Exception:
                continue

        return records


# ─── 建设银行 ───────────────────────────────────────────────────────────
class CCBAdapter(BankStatementAdapter):
    """
    建设银行对账单格式：
    通常 Row1 含 "中国建设银行" 或 "CCB"
    表头包含：交易日期|交易时间|交易金额|余额|对方账户|对方户名|摘要 等
    """
    bank_code = 'CCB'
    bank_name = '建设银行'

    def detect(self, ws) -> bool:
        try:
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = str(ws.cell(row, col).value or '')
                    if '建设银行' in val or 'CCB' in val.upper():
                        return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        # 找表头行
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1)]
            if any('日期' in v and '金额' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            header_row = 1

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    return ws.cell(row, c).value
            return None

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                # 尝试找金额列（收入或支出）
                income_val = self._decimal(get_col(row_idx, '贷方金额', '收入金额', '存入金额'))
                expense_val = self._decimal(get_col(row_idx, '借方金额', '支出金额', '支取金额'))

                if income_val is not None and income_val > 0:
                    amount, direction = income_val, 'income'
                elif expense_val is not None and expense_val > 0:
                    amount, direction = expense_val, 'expense'
                else:
                    # 尝试从余额变化推导
                    amount_raw = get_col(row_idx, '交易金额', '金额')
                    amt = self._decimal(amount_raw)
                    if amt is None or amt == 0:
                        continue
                    amount, direction = abs(amt), 'income' if str(amount_raw).strip().startswith('+') else 'expense'

                txn_date = self._parse_date(get_col(row_idx, '交易日期', '记账日期', '日期'))
                if not txn_date:
                    continue

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=self._parse_time(get_col(row_idx, '交易时间', '时间')),
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '余额')),
                    counterparty_name=str(get_col(row_idx, '对方户名', '收款人', '付款人') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '对方账号', '收款账号', '付款账号') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '对方银行', '开户行') or '').strip(),
                    summary=str(get_col(row_idx, '摘要', '交易描述', '说明') or '').strip(),
                    bank_serial=str(get_col(row_idx, '流水号', '交易流水') or '').strip(),
                    raw_data=dict(zip(headers, [ws.cell(row_idx, c).value for c in range(1, len(headers) + 1)]))
                ))
            except Exception:
                continue

        return records


# ─── 中国银行 ───────────────────────────────────────────────────────────
class BOCAdapter(BankStatementAdapter):
    """
    中国银行对账单格式：
    通常 Row1 含 "中国银行" 或 "BOC"
    """
    bank_code = 'BOC'
    bank_name = '中国银行'

    def detect(self, ws) -> bool:
        try:
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = str(ws.cell(row, col).value or '')
                    if '中国银行' in val or 'BOC' in val.upper():
                        return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1)]
            if any('日期' in v and '金额' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            header_row = 1

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    return ws.cell(row, c).value
            return None

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                income_val = self._decimal(get_col(row_idx, '贷方发生额', '收入金额', '存入金额'))
                expense_val = self._decimal(get_col(row_idx, '借方发生额', '支出金额', '支取金额'))

                if income_val is not None and income_val > 0:
                    amount, direction = income_val, 'income'
                elif expense_val is not None and expense_val > 0:
                    amount, direction = expense_val, 'expense'
                else:
                    continue

                txn_date = self._parse_date(get_col(row_idx, '交易日期', '记账日期', '日期'))
                if not txn_date:
                    continue

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=self._parse_time(get_col(row_idx, '交易时间')),
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '账户余额', '余额')),
                    counterparty_name=str(get_col(row_idx, '对方名称', '对方户名', '收款人', '付款人') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '对方账号', '对方账户') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '对方开户行', '对方银行') or '').strip(),
                    summary=str(get_col(row_idx, '摘要', '用途', '交易描述') or '').strip(),
                    bank_serial=str(get_col(row_idx, '流水号', '交易流水号', '参考号') or '').strip(),
                    raw_data=dict(zip(headers, [ws.cell(row_idx, c).value for c in range(1, len(headers) + 1)]))
                ))
            except Exception:
                continue

        return records


# ─── 农业银行 ───────────────────────────────────────────────────────────
class ABCAdapter(BankStatementAdapter):
    """
    农业银行对账单格式
    """
    bank_code = 'ABC'
    bank_name = '农业银行'

    def detect(self, ws) -> bool:
        try:
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = str(ws.cell(row, col).value or '')
                    if '农业银行' in val or 'ABC' in val.upper():
                        return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1)]
            if any('日期' in v and '金额' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            header_row = 1

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    return ws.cell(row, c).value
            return None

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                income_val = self._decimal(get_col(row_idx, '贷方金额', '收入金额', '存入金额'))
                expense_val = self._decimal(get_col(row_idx, '借方金额', '支出金额', '支取金额'))

                if income_val is not None and income_val > 0:
                    amount, direction = income_val, 'income'
                elif expense_val is not None and expense_val > 0:
                    amount, direction = expense_val, 'expense'
                else:
                    continue

                txn_date = self._parse_date(get_col(row_idx, '交易日期', '记账日期', '日期'))
                if not txn_date:
                    continue

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=self._parse_time(get_col(row_idx, '交易时间')),
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '账户余额', '余额')),
                    counterparty_name=str(get_col(row_idx, '对方名称', '对方户名', '收款人', '付款人') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '对方账号', '对方账户') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '对方开户行', '对方银行') or '').strip(),
                    summary=str(get_col(row_idx, '摘要', '用途', '交易描述') or '').strip(),
                    bank_serial=str(get_col(row_idx, '流水号', '交易流水号') or '').strip(),
                    raw_data=dict(zip(headers, [ws.cell(row_idx, c).value for c in range(1, len(headers) + 1)]))
                ))
            except Exception:
                continue

        return records


# ─── 交通银行 ───────────────────────────────────────────────────────────
class COMMAdapter(BankStatementAdapter):
    """
    交通银行对账单格式
    """
    bank_code = 'COMM'
    bank_name = '交通银行'

    def detect(self, ws) -> bool:
        try:
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = str(ws.cell(row, col).value or '')
                    if '交通银行' in val or 'COMM' in val.upper():
                        return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1)]
            if any('日期' in v and '金额' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            header_row = 1

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    return ws.cell(row, c).value
            return None

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                income_val = self._decimal(get_col(row_idx, '贷方金额', '收入金额', '存入金额'))
                expense_val = self._decimal(get_col(row_idx, '借方金额', '支出金额', '支取金额'))

                if income_val is not None and income_val > 0:
                    amount, direction = income_val, 'income'
                elif expense_val is not None and expense_val > 0:
                    amount, direction = expense_val, 'expense'
                else:
                    continue

                txn_date = self._parse_date(get_col(row_idx, '交易日期', '记账日期', '日期'))
                if not txn_date:
                    continue

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=self._parse_time(get_col(row_idx, '交易时间')),
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '账户余额', '余额')),
                    counterparty_name=str(get_col(row_idx, '对方名称', '对方户名', '收款人', '付款人') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '对方账号', '对方账户') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '对方开户行', '对方银行') or '').strip(),
                    summary=str(get_col(row_idx, '摘要', '用途', '交易描述') or '').strip(),
                    bank_serial=str(get_col(row_idx, '流水号', '交易流水号') or '').strip(),
                    raw_data=dict(zip(headers, [ws.cell(row_idx, c).value for c in range(1, len(headers) + 1)]))
                ))
            except Exception:
                continue

        return records


# ─── 邮储银行 ───────────────────────────────────────────────────────────
class PSBCAdapter(BankStatementAdapter):
    """
    中国邮政储蓄银行对账单格式
    """
    bank_code = 'PSBC'
    bank_name = '邮储银行'

    def detect(self, ws) -> bool:
        try:
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = str(ws.cell(row, col).value or '')
                    if '邮储' in val or '邮政储蓄' in val or 'PSBC' in val.upper():
                        return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1)]
            if any('日期' in v and '金额' in v for v in row_vals):
                header_row = r
                break
        if header_row is None:
            header_row = 1

        headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    return ws.cell(row, c).value
            return None

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                income_val = self._decimal(get_col(row_idx, '贷方金额', '收入金额', '存入金额'))
                expense_val = self._decimal(get_col(row_idx, '借方金额', '支出金额', '支取金额'))

                if income_val is not None and income_val > 0:
                    amount, direction = income_val, 'income'
                elif expense_val is not None and expense_val > 0:
                    amount, direction = expense_val, 'expense'
                else:
                    continue

                txn_date = self._parse_date(get_col(row_idx, '交易日期', '记账日期', '日期'))
                if not txn_date:
                    continue

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=self._parse_time(get_col(row_idx, '交易时间')),
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '账户余额', '余额')),
                    counterparty_name=str(get_col(row_idx, '对方名称', '对方户名', '收款人', '付款人') or '').strip(),
                    counterparty_account=str(get_col(row_idx, '对方账号', '对方账户') or '').strip(),
                    counterparty_bank=str(get_col(row_idx, '对方开户行', '对方银行') or '').strip(),
                    summary=str(get_col(row_idx, '摘要', '用途', '交易描述') or '').strip(),
                    bank_serial=str(get_col(row_idx, '流水号', '交易流水号') or '').strip(),
                    raw_data=dict(zip(headers, [ws.cell(row_idx, c).value for c in range(1, len(headers) + 1)]))
                ))
            except Exception:
                continue

        return records


class XlrdSheetWrapper:
    """将 xlrd sheet 适配为类似 openpyxl 的工作表接口，供各银行适配器使用。"""
    def __init__(self, sheet):
        self.sheet = sheet
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row: int, col: int):
        """openpyxl 风格：row/col 从 1 开始"""
        return XlrdCell(self.sheet.cell_value(row - 1, col - 1))


class XlrdCell:
    def __init__(self, value):
        self.value = value

# ─── 平安银行 ───────────────────────────────────────────────────────────
class PingAnAdapter(BankStatementAdapter):
    """
    平安银行对账单格式（2026-05实测）：
    Row1 (A1): ''（空行，sheet名=账号）
    Row2: 表头行，含11列：
      交易日期 | 账号 | 借 | 贷 | 账户余额 | 对方账户 | 对方账户名称 | 交易流水号 | 单位结算卡号 | 摘要 | 用途
    Row3+: 数据行

    特点：
    - 借贷分列：借/贷两列互斥（有借无贷，有贷无借）
    - 我方账号在Row1的sheet名，或第2列
    - 流水号（交易流水号）唯一性高，适合去重
    - 摘要是交易渠道（网银/跨行转账等），用途是业务内容（货款/运费等）
    """
    bank_code = 'PINGAN'
    bank_name = '平安银行'

    def detect(self, ws) -> bool:
        """
        平安银行专属标志：账号列以 '1500' 开头。
        检测前10行数据行，只要有一行账号符合即匹配。
        """
        # openpyxl cell() requires 1-indexed rows (row >= 1), but ws.max_row is 1-indexed
        # so r in range must be offset by +1 when calling ws.cell()
        try:
            # 先确认表头存在（任意行）
            header_found = False
            for r in range(min(5, ws.max_row)):
                # r is 0-indexed here; ws.cell() needs 1-indexed row
                headers = [str(ws.cell(r + 1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
                if any(h == '账号' for h in headers):
                    header_found = True
                    break
            if not header_found:
                return False
            # 用账号列的值（以1500开头）作为专属标志
            for r in range(min(10, ws.max_row)):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(r + 1, c).value or '').strip()
                    if val.startswith('1500') and len(val) >= 8:
                        return True
            return False
        except Exception:
            return False

    def parse(self, ws) -> list[ParsedTransaction]:
        # 动态查找表头行（平安银行文件表头在任意行）
        # 注意：ws.cell() requires 1-indexed rows; range indices are 0-indexed
        header_row = None
        for r in range(min(10, ws.max_row)):
            headers = [str(ws.cell(r + 1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if any(h in headers for h in ['交易日期', '账号', '借', '贷', '账户余额']):
                header_row = r  # 0-indexed
                break
        if header_row is None:
            return []
        headers = [str(ws.cell(header_row + 1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        col_idx = {h: c for c, h in enumerate(headers, start=1)}

        def get_col(row, *names):
            for name in names:
                c = col_idx.get(name)
                if c:
                    val = ws.cell(row + 1, c).value  # row is 0-indexed, +1 for ws.cell()
                    if val is not None:
                        return val
            return None

        records = []
        # 提取我方账号（从第一个有效数据行的"账号"列）
        file_account_no = ''
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                debit_raw  = get_col(row_idx, '借')
                credit_raw = get_col(row_idx, '贷')
                debit  = self._decimal(debit_raw)
                credit = self._decimal(credit_raw)
                if (debit is not None and debit > 0) or (credit is not None and credit > 0):
                    file_account_no = str(get_col(row_idx, '账号') or '').strip()
                    break
            except Exception:
                continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                debit_raw  = get_col(row_idx, '借')
                credit_raw = get_col(row_idx, '贷')

                debit  = self._decimal(debit_raw)
                credit = self._decimal(credit_raw)

                if debit is not None and debit > 0:
                    amount, direction = debit, 'expense'
                elif credit is not None and credit > 0:
                    amount, direction = credit, 'income'
                else:
                    continue

                txn_date = self._parse_date(get_col(row_idx, '交易日期'))
                if not txn_date:
                    continue

                cp_name    = str(get_col(row_idx, '对方账户名称') or '').strip()
                cp_account = str(get_col(row_idx, '对方账户') or '').strip()
                cp_bank    = ''
                summary    = str(get_col(row_idx, '摘要') or '').strip()

                records.append(ParsedTransaction(
                    transaction_date=txn_date,
                    transaction_time=None,
                    amount=amount,
                    direction=direction,
                    balance=self._decimal(get_col(row_idx, '账户余额')),
                    bank_serial=str(get_col(row_idx, '交易流水号') or '').strip(),
                    counterparty_name=cp_name,
                    counterparty_account=cp_account,
                    counterparty_bank=cp_bank,
                    summary=summary,
                    transaction_type=str(get_col(row_idx, '用途') or '').strip(),
                    account_no=file_account_no,
                ))
            except Exception:
                continue

        return records


# ─── 自动识别解析 ───────────────────────────────────────────────────────
def detect_and_parse(file_content):
    """自动识别格式并解析，返回 (bank_code, transactions)
    支持 .xlsx (openpyxl) 和 .xls (xlrd) 两种格式。
    file_content 可以是 bytes（文件路径读取结果）或 BytesIO（Django InMemoryUploadedFile.read() 返回值）。
    """
    import io
    if isinstance(file_content, (io.IOBase,)):
        file_obj = file_content
    else:
        file_obj = io.BytesIO(file_content)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception:
        import xlrd
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        xlrd_content = file_obj.read()
        wb = xlrd.open_workbook(file_contents=xlrd_content)
        ws = XlrdSheetWrapper(wb.sheet_by_index(0))

    for adapter_cls in ALL_ADAPTERS:
        adapter = adapter_cls()
        if adapter.detect(ws):
            txns = adapter.parse(ws)
            # 统一补 account_no：若所有 txn.account_no 为空，尝试从第1列账号提取
            if txns and all(not t.account_no for t in txns):
                import re as _re
                _acc = ''
                for r in range(1, min(ws.max_row + 1, 50)):
                    for c in range(1, ws.max_column + 1):
                        v = str(ws.cell(r, c).value or '').strip()
                        if _re.fullmatch(r'\d{6,20}', v):
                            _acc = v
                            break
                    if _acc:
                        break
                if _acc:
                    for t in txns:
                        t.account_no = _acc
            return adapter_cls.bank_code, txns

    raise ValueError("无法识别银行格式，请确认文件为以下银行对账单：工商银行、招商银行、建设银行、中国银行、农业银行、交通银行、邮储银行、平安银行")


def parse_with_adapter(file_content, bank_code: str):
    """用指定银行适配器解析，支持 .xlsx / .xls"""
    # 数据库中平安银行用 PA，前端模板用 PINGAN，统一标准化
    bank_code = 'PINGAN' if bank_code == 'PA' else bank_code
    adapters = {a.bank_code: a for a in [cls() for cls in ALL_ADAPTERS]}
    if bank_code not in adapters:
        raise ValueError(f"不支持的银行: {bank_code}，支持的银行：{', '.join(adapters.keys())}")

    import io
    if isinstance(file_content, (io.IOBase,)):
        file_obj = file_content
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
    else:
        file_obj = io.BytesIO(file_content)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception:
        import xlrd
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        xlrd_content = file_obj.read()
        wb = xlrd.open_workbook(file_contents=xlrd_content)
        ws = XlrdSheetWrapper(wb.sheet_by_index(0))

    return adapters[bank_code].parse(ws)


def detect_with_adapter(file_content, bank_code: str) -> bool:
    """
    用指定银行的适配器检测文件格式是否匹配。
    返回 True=匹配，False=不匹配。
    不会抛出异常。
    """
    bank_code = 'PINGAN' if bank_code == 'PA' else bank_code
    adapters = {a.bank_code: a for a in [cls() for cls in ALL_ADAPTERS]}
    if bank_code not in adapters:
        return False

    import io
    if isinstance(file_content, (io.IOBase,)):
        file_obj = file_content
    else:
        file_obj = io.BytesIO(file_content)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception:
        try:
            import xlrd
            if hasattr(file_obj, 'seek'):
                file_obj.seek(0)
            xlrd_content = file_obj.read()
            wb = xlrd.open_workbook(file_contents=xlrd_content)
            ws = XlrdSheetWrapper(wb.sheet_by_index(0))
        except Exception:
            return False

    return adapters[bank_code].detect(ws)


# ─── 适配器注册表（CMB在前，更具体优先检测）─────────────────────────────
ALL_ADAPTERS = [
    CMBAdapter,     # 招商银行（最具体，优先）
    ICBCAdapter,    # 工商银行
    CCBAdapter,     # 建设银行
    BOCAdapter,     # 中国银行
    ABCAdapter,     # 农业银行
    COMMAdapter,    # 交通银行
    PSBCAdapter,    # 邮储银行
    PingAnAdapter,  # 平安银行
]
