"""
Excel批量导入视图 - CRM模块
POST /api/crm/import/clients/     - 客户导入
POST /api/crm/import/suppliers/   - 供应商导入
POST /api/crm/import/contracts/  - 合同导入
"""
import datetime
import io
import re
from decimal import Decimal

from django.db import transaction
from openpyxl import load_workbook
from apps.core.permissions import require_perms
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from apps.crm.models import Client, Supplier, Contract


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.date() if isinstance(value, datetime.datetime) else value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y'):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_decimal(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value).replace(',', '').replace('¥', '').replace(' ', '').strip())
    except (ValueError, TypeError):
        return None


def _cell(ws, row_num: int, col_idx: int):
    if col_idx is None:
        return None
    try:
        return ws[row_num][col_idx].value
    except IndexError:
        return None


class ImportResult:
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors = []
        self.created_ids = []

    def add_error(self, row, field, msg, value=None):
        self.errors.append({
            'row': row, 'field': field, 'message': msg,
            'value': str(value) if value is not None else None
        })
        self.error_count += 1

    def to_dict(self):
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'errors': self.errors,
            'created_ids': self.created_ids,
        }


def _col_map(ws, mapping: dict[str, str]) -> dict[str, int]:
    headers = [c.value for c in ws[1]]
    return {f: headers.index(h) for f, h in mapping.items() if h in headers}


# ─── 客户导入 ────────────────────────────────────────────────

CLIENT_FIELDS = {
    'name': '客户名称',
    'category': '客户类别',
    'contact_person': '联系人',
    'contact_phone': '联系电话',
    'contact_email': '邮箱',
    'address': '地址',
    'remark': '备注',
}


@api_view(['POST'])
@require_perms('crm:client:create')
def import_clients(request):
    """导入客户。必填：客户名称"""
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    cm = _col_map(ws, CLIENT_FIELDS)
    if 'name' not in cm:
        return Response({'error': '缺少必需列：客户名称'}, status=400)

    result = ImportResult()
    for row_num in range(2, ws.max_row + 1):
        name = str(_cell(ws, row_num, cm['name']) or '').strip()
        if not name:
            continue

        try:
            with transaction.atomic():
                obj = Client.objects.create(
                    name=name,
                    category=str(_cell(ws, row_num, cm.get('category')) or '企业客户').strip(),
                    contact_person=str(_cell(ws, row_num, cm.get('contact_person')) or '').strip(),
                    contact_phone=str(_cell(ws, row_num, cm.get('contact_phone')) or '').strip(),
                    contact_email=str(_cell(ws, row_num, cm.get('contact_email')) or '').strip(),
                    address=str(_cell(ws, row_num, cm.get('address')) or '').strip(),
                    remark=str(_cell(ws, row_num, cm.get('remark')) or '').strip(),
                    created_by=request.user,
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())


# ─── 供应商导入 ─────────────────────────────────────────────

SUPPLIER_FIELDS = {
    'name': '供应商名称',
    'contact_person': '联系人',
    'contact_phone': '联系电话',
    'contact_email': '邮箱',
    'brands': '代理品牌',
    'address': '地址',
    'remark': '备注',
}


@api_view(['POST'])
@require_perms('crm:supplier:create')
def import_suppliers(request):
    """导入供应商。必填：供应商名称"""
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    cm = _col_map(ws, SUPPLIER_FIELDS)
    if 'name' not in cm:
        return Response({'error': '缺少必需列：供应商名称'}, status=400)

    result = ImportResult()
    for row_num in range(2, ws.max_row + 1):
        name = str(_cell(ws, row_num, cm['name']) or '').strip()
        if not name:
            continue

        try:
            with transaction.atomic():
                obj = Supplier.objects.create(
                    name=name,
                    contact_person=str(_cell(ws, row_num, cm.get('contact_person')) or '').strip(),
                    contact_phone=str(_cell(ws, row_num, cm.get('contact_phone')) or '').strip(),
                    contact_email=str(_cell(ws, row_num, cm.get('contact_email')) or '').strip(),
                    brands=str(_cell(ws, row_num, cm.get('brands')) or '').strip(),
                    address=str(_cell(ws, row_num, cm.get('address')) or '').strip(),
                    remark=str(_cell(ws, row_num, cm.get('remark')) or '').strip(),
                    status='active',
                    created_by=request.user,
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())


# ─── 合同导入 ────────────────────────────────────────────────

CONTRACT_FIELDS = {
    'contract_no': '合同编号',
    'name': '合同名称',
    'counterparty_type': '对方类型',
    'amount': '金额',
    'sign_date': '签订日期',
    'expire_date': '到期日期',
    'status': '状态',
    'remark': '备注',
}


@api_view(['POST'])
@require_perms('crm:contract:create')
def import_contracts(request):
    """导入合同。必填：合同编号、合同名称、合同金额、签订日期"""
    if 'file' not in request.FILES:
        return Response({'error': '请上传 .xlsx 文件'}, status=400)

    wb = load_workbook(io.BytesIO(request.FILES['file'].read()), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return Response({'error': '文件无数据行'}, status=400)

    headers = [c.value for c in ws[1]]
    cm = {f: headers.index(h) for f, h in CONTRACT_FIELDS.items() if h in headers}
    # 额外列：客户名称/供应商名称（用于查找FK）
    client_col = headers.index('客户名称') if '客户名称' in headers else None
    supplier_col = headers.index('供应商名称') if '供应商名称' in headers else None
    project_col = headers.index('项目名称') if '项目名称' in headers else None

    result = ImportResult()
    COUNTERPARTY_TYPE_MAP = {'客户': 'client', '供应商': 'supplier'}

    for row_num in range(2, ws.max_row + 1):
        contract_no = str(_cell(ws, row_num, cm.get('contract_no')) or '').strip()
        name = str(_cell(ws, row_num, cm.get('name')) or '').strip()
        if not contract_no or not name:
            result.add_error(row_num, 'general', '合同编号和合同名称不能为空')
            continue

        amount = _parse_decimal(_cell(ws, row_num, cm.get('amount')))
        if amount is None:
            result.add_error(row_num, '合同金额', '金额格式错误')
            continue

        sign_date = _parse_date(_cell(ws, row_num, cm.get('sign_date')))
        if sign_date is None:
            result.add_error(row_num, '签订日期', '日期格式错误')
            continue

        expire_date = _parse_date(_cell(ws, row_num, cm.get('expire_date')))

        counterparty_type_val = COUNTERPARTY_TYPE_MAP.get(
            str(_cell(ws, row_num, cm.get('counterparty_type')) or '客户').strip(), 'client')

        client = None
        supplier = None
        if counterparty_type_val == 'client' and client_col is not None:
            client_name = str(_cell(ws, row_num, client_col) or '').strip()
            if client_name:
                client = Client.objects.filter(name=client_name).first()
        elif counterparty_type_val == 'supplier' and supplier_col is not None:
            supplier_name = str(_cell(ws, row_num, supplier_col) or '').strip()
            if supplier_name:
                supplier = Supplier.objects.filter(name=supplier_name).first()

        project = None
        if project_col is not None:
            project_name = str(_cell(ws, row_num, project_col) or '').strip()
            if project_name:
                from apps.tasks.models import Project
                project = Project.objects.filter(name=project_name).first()

        try:
            with transaction.atomic():
                obj = Contract.objects.create(
                    contract_no=contract_no,
                    name=name,
                    counterparty_type=counterparty_type_val,
                    client=client,
                    supplier=supplier,
                    project=project,
                    amount=amount,
                    sign_date=sign_date,
                    expire_date=expire_date,
                    status=str(_cell(ws, row_num, cm.get('status')) or 'draft').strip(),
                    remark=str(_cell(ws, row_num, cm.get('remark')) or '').strip(),
                    created_by=request.user,
                )
                result.created_ids.append(obj.id)
                result.success_count += 1
        except Exception as ex:
            result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

    return Response(result.to_dict())
