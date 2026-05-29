"""
Excel批量导入引擎 - 基于openpyxl
支持：字段映射、逐行校验、批量写入、错误报告
"""

import io
from decimal import Decimal
from typing import Any, Type

from django.db import transaction
from openpyxl import load_workbook


class ImportError:
    """单行错误"""

    def __init__(self, row: int, field: str, message: str, value: Any = None):
        self.row = row
        self.field = field
        self.message = message
        self.value = value

    def as_dict(self):
        return {
            'row': self.row,
            'field': self.field,
            'message': self.message,
            'value': str(self.value) if self.value is not None else None,
        }


class ImportResult:
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors: list[ImportError] = []
        self.created_ids: list[int] = []

    def add_error(self, row: int, field: str, message: str, value: Any = None):
        self.errors.append(ImportError(row, field, message, value))
        self.error_count += 1

    def as_dict(self):
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'errors': [e.as_dict() for e in self.errors],
            'created_ids': self.created_ids,
        }


class ExcelImporter:
    """
    通用Excel导入器。

    用法示例：

        class EmployeeImporter(ExcelImporter):
            model = Employee
            required_fields = ['name', 'code', 'company']
            field_mapping = {
                '姓名': 'name',
                '工号': 'code',
                '所属公司': 'company',
                '手机': 'phone',
            }
            number_fields = ['phone']
            date_fields = ['hire_date']

            def clean_value(self, field: str, value: Any) -> Any:
                # 自定义清洗逻辑
                if field == 'phone':
                    return re.sub(r'\\D', '', str(value))
                return value
    """

    model: Type = None
    required_fields: list[str] = []
    # Excel表头名 → model字段名
    field_mapping: dict[str, str] = {}
    number_fields: list[str] = []  # 需要转数值的字段
    decimal_fields: list[str] = []  # 需要转Decimal的字段
    date_fields: list[str] = []  # 需要转date的字段
    boolean_fields: list[str] = []  # 需要转bool的字段

    def __init__(self, file_content: bytes):
        self.wb = load_workbook(io.BytesIO(file_content), data_only=True)
        self.ws = self.wb.active
        self.result = ImportResult()
        # 反向映射：model字段 → Excel列索引
        self._col_map: dict[str, int] = {}

    def _parse_headers(self):
        """解析表头行，建立列索引"""
        headers = []
        for cell in self.ws[1]:
            headers.append(cell.value)
        for model_field, excel_header in self.field_mapping.items():
            if excel_header in headers:
                self._col_map[model_field] = headers.index(excel_header)

    def _get_cell_value(self, row_num: int, model_field: str) -> tuple[Any, str]:
        """
        获取指定行指定字段的值。
        Returns (raw_value, excel_header_name)
        """
        col_idx = self._col_map.get(model_field)
        if col_idx is None:
            return None, ''
        cell = self.ws[row_num][col_idx]
        return cell.value, self._find_header_by_col(col_idx)

    def _find_header_by_col(self, col_idx: int) -> str:
        for model_field, excel_header in self.field_mapping.items():
            if self._col_map.get(model_field) == col_idx:
                return excel_header
        return ''

    def clean_value(self, field: str, value: Any) -> Any:
        """子类可覆盖：自定义字段清洗逻辑"""
        if value is None or value == '':
            return None
        if field in self.number_fields:
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return None
        if field in self.decimal_fields:
            try:
                return Decimal(str(value).replace(',', '').replace('¥', '').strip())
            except (ValueError, TypeError):
                return None
        if field in self.boolean_fields:
            if isinstance(value, bool):
                return value
            v = str(value).strip().lower()
            return v in ('是', 'yes', 'true', '1', '有', '已认证')
        return value

    def validate(self, row_num: int, row_data: dict[str, Any]) -> list[ImportError]:
        """子类可覆盖：返回错误列表"""
        errors = []
        for field in self.required_fields:
            if not row_data.get(field):
                errors.append(ImportError(row_num, field, f'必填字段 [{field}] 不能为空'))
        return errors

    def get_company_id(self, row_data: dict) -> int | None:
        """子类可覆盖：从row_data或company字段获取company_id"""
        return row_data.get('company')

    def save_row(self, row_data: dict[str, Any]) -> object:
        """子类可覆盖：自定义单行保存逻辑"""
        return self.model.objects.create(**row_data)

    def run(self) -> ImportResult:
        """执行导入，返回结果"""
        self._parse_headers()
        rows = list(self.ws.iter_rows(min_row=2, values_only=True))

        # 快速校验：检查必填字段是否都有映射
        missing = []
        for field in self.required_fields:
            if field not in self._col_map:
                missing.append(field)
        if missing:
            self.result.add_error(0, '表头', f'缺少必需列：{", ".join(missing)}')
            return self.result

        # 逐行处理
        for row_num, row in enumerate(rows, start=2):
            if all(cell is None for cell in row):
                continue  # 跳过空行

            row_data = {}
            for model_field, col_idx in self._col_map.items():
                if col_idx < len(row):
                    raw = row[col_idx]
                    cleaned = self.clean_value(model_field, raw)
                    if cleaned is not None:
                        row_data[model_field] = cleaned

            # 校验
            errors = self.validate(row_num, row_data)
            if errors:
                for e in errors:
                    self.result.add_error(e.row, e.field, e.message, e.value)
                continue

            try:
                with transaction.atomic():
                    obj = self.save_row(row_data)
                    self.result.created_ids.append(obj.id)
                    self.result.success_count += 1
            except Exception as ex:
                self.result.add_error(row_num, 'general', f'保存失败: {str(ex)}')

        return self.result
