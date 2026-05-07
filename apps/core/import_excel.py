"""
Excel 批量导入工具 — 使用 openpyxl 解析导入文件
支持：收入/支出/发票/工资 四个模块的批量导入
"""
import io
import datetime
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation


class ImportResult:
    """导入结果封装"""
    def __init__(self):
        self.success = 0
        self.error = 0
        self.errors = []  # [{row, message}]
        self.rows = []    # 解析后的行数据字典

    def add_error(self, row_num, msg):
        self.error += 1
        self.errors.append({'row': row_num, 'message': msg})

    def to_dict(self):
        return {
            'success': self.success,
            'error': self.error,
            'errors': self.errors[:20],  # 最多返回20条错误
            'rows': self.rows
        }


def parse_uploaded_file(file_obj, expected_headers=None):
    """
    通用 Excel 解析器。
    file_obj: Django UploadedFile 或 BytesIO
    expected_headers: 期望的表头列表（子串匹配即可）
    返回 (headers, all_rows)
    """
    if hasattr(file_obj, 'read'):
        data = file_obj.read()
    else:
        data = file_obj



    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    # 跳过空行，找到第一行非空作为表头
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell is not None for cell in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError('Excel 文件为空或无有效数据')

    # 读取表头
    header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    headers = [str(h).strip() if h is not None else '' for h in header_row]

    # 读取数据行
    all_rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(cell is None for cell in row):
            continue
        all_rows.append(list(row))

    return headers, all_rows


def match_header(headers, *keywords):
    """在 headers 中查找包含任意 keyword 的列索引（0-based）。找不到返回 -1。"""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in str(h):
                return i
    return -1


def parse_number(val):
    """把各种格式的数字字符串转 Decimal，失败返回 None。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '').replace('¥', '').replace('元', '').replace(' ', '')
    if not s or s == '-':
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_date(val):
    """解析日期，返回 'YYYY-MM-DD' 字符串或 None。"""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.date):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # 尝试多种格式
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


def parse_int(val):
    """解析整数。"""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


# ─── 收入导入 ────────────────────────────────────────────────────────

def import_income(file_obj, company_id=None, operator=None):
    """
    导入收入记录。
    必需列：金额, 日期
    可选列：公司(名称), 项目(名称), 来源, 状态, 备注
    company_id: 当前用户所在公司（作为默认值）
    """
    result = ImportResult()
    headers, all_rows = parse_uploaded_file(file_obj)

    # 找列索引（1-based）
    col_amount   = match_header(headers, '金额', 'amount', '收入金额')
    col_date     = match_header(headers, '日期', 'date', '收入日期')
    col_company  = match_header(headers, '公司', 'company', '所属公司')
    col_project  = match_header(headers, '项目', 'project', '所属项目')
    col_source   = match_header(headers, '来源', 'source', '收入来源', '类别')
    col_status   = match_header(headers, '状态', 'status')
    col_desc     = match_header(headers, '备注', 'description', '描述', '说明')
    col_operator = match_header(headers, '操作人', 'operator', '经办人')

    if col_amount == -1:
        result.add_error(0, '未找到"金额"列，请确保 Excel 包含金额列')
        return result
    if col_date == -1:
        result.add_error(0, '未找到"日期"列，请确保 Excel 包含日期列')
        return result

    from apps.finance.models import Income, Company
    from apps.tasks.models import Project

    # 缓存
    company_cache = {c.name: c.id for c in Company.objects.all()}
    project_cache = {p.name: p.id for p in Project.objects.all()}
    # 收入状态映射（confirmed → pending，保证非法值不写入 DB）
    status_map = {'approved': 'approved', '已批准': 'approved', '待审批': 'pending',
                  'pending': 'pending', 'rejected': 'rejected', '已拒绝': 'rejected',
                  'confirmed': 'pending', '已确认': 'pending', '草稿': 'draft', 'draft': 'draft'}

    for i, row in enumerate(all_rows, start=2):
        try:
            amount = parse_number(row[col_amount] if col_amount >= 0 and col_amount < len(row) else None)
            date   = parse_date(row[col_date] if col_date >= 0 and col_date < len(row) else None)

            if not amount:
                result.add_error(i, f'行{i}：金额无效')
                continue
            if not date:
                result.add_error(i, f'行{i}：日期无效')
                continue

            company_name = row[col_company].strip() if col_company >= 0 and col_company < len(row) and row[col_company] else None
            project_name = row[col_project].strip() if col_project >= 0 and col_project < len(row) and row[col_project] else None
            source = row[col_source] if col_source >= 0 and col_source < len(row) else None
            if source:
                source = str(source).strip()
            status_val = row[col_status] if col_status >= 0 and col_status < len(row) else None
            # Income 状态: pending/approved/rejected，无 confirmed（已改 approved）
            status = status_map.get(str(status_val).strip(), 'pending') if status_val else 'pending'
            desc = row[col_desc] if col_desc >= 0 and col_desc < len(row) else None
            if desc:
                desc = str(desc).strip()

            company_id_val = None
            project_id_val = None
            if company_name:
                company_id_val = company_cache.get(company_name)
            if project_name:
                project_id_val = project_cache.get(project_name)

            result.rows.append({
                'company': company_id_val,
                'company_name': company_name,
                'project': project_id_val,
                'project_name': project_name,
                'source': source or '',
                'amount': float(amount),
                'date': date,
                'status': status,
                'description': desc or '',
            })
            result.success += 1
        except Exception as e:
            result.add_error(i, f'行{i} 解析异常：{str(e)}')

    return result


# ─── 支出导入 ────────────────────────────────────────────────────────

def import_expense(file_obj, company_id=None, operator=None):
    """
    导入支出记录。
    必需列：金额, 日期
    可选列：公司, 支出类型, 项目, 付款方式, 状态, 供应商/收款方, 备注
    """
    result = ImportResult()
    headers, all_rows = parse_uploaded_file(file_obj)

    col_amount   = match_header(headers, '金额', 'amount', '支出金额')
    col_date     = match_header(headers, '日期', 'date', '支出日期')
    col_company  = match_header(headers, '公司', 'company', '所属公司')
    col_type     = match_header(headers, '类型', 'type', '支出类型', '费用类型')
    col_project  = match_header(headers, '项目', 'project', '所属项目')
    col_payee    = match_header(headers, '供应商', '收款方', 'payee', '供应商/收款方')
    # 注意：Expense 模型没有 payment_method/payee 字段，已移除对应导入逻辑
    col_status   = match_header(headers, '状态', 'status')
    col_desc     = match_header(headers, '备注', 'description', '描述', '说明')

    if col_amount == -1:
        result.add_error(0, '未找到"金额"列')
        return result
    if col_date == -1:
        result.add_error(0, '未找到"日期"列')
        return result

    from apps.finance.models import Expense, Company
    from apps.tasks.models import Project

    company_cache = {c.name: c.id for c in Company.objects.all()}
    project_cache = {p.name: p.id for p in Project.objects.all()}

    # 支出类型映射（双向：英文值→英文值 + 中文标签→英文值）
    raw_choices = dict(Expense.EXPENSE_TYPE_CHOICES) if hasattr(Expense, 'EXPENSE_TYPE_CHOICES') else {}
    # 正向: expense→expense, advance→advance; 反向: 费用报销→expense, 预付款→advance
    type_choices = {}
    for k, label in raw_choices.items():
        type_choices[k] = k          # expense→expense
        type_choices[label] = k      # 费用报销→expense
    status_map = {'approved': 'approved', '已批准': 'approved', '待审批': 'pending', 'pending': 'pending', 'draft': 'draft', '草稿': 'draft', 'rejected': 'rejected', '已拒绝': 'rejected'}

    def normalize_choice(val, mapping):
        if not val:
            return None
        v = str(val).strip()
        return mapping.get(v, v)

    for i, row in enumerate(all_rows, start=2):
        try:
            amount = parse_number(row[col_amount] if col_amount >= 0 and col_amount < len(row) else None)
            date   = parse_date(row[col_date] if col_date >= 0 and col_date < len(row) else None)

            if not amount:
                result.add_error(i, f'行{i}：金额无效')
                continue
            if not date:
                result.add_error(i, f'行{i}：日期无效')
                continue

            company_name = row[col_company] if col_company >= 0 and col_company < len(row) else None
            company_name = company_name.strip() if company_name else None
            project_name = row[col_project] if col_project >= 0 and col_project < len(row) else None
            project_name = project_name.strip() if project_name else None
            type_val     = row[col_type] if col_type >= 0 and col_type < len(row) else None
            payee_val    = row[col_payee] if col_payee >= 0 and col_payee < len(row) else None
            status_val   = row[col_status] if col_status >= 0 and col_status < len(row) else None
            desc_val     = row[col_desc] if col_desc >= 0 and col_desc < len(row) else None

            result.rows.append({
                'company': company_cache.get(company_name) if company_name else None,
                'company_name': company_name,
                'expense_type': normalize_choice(type_val, type_choices) if 'type_choices' in dir() else (type_val or 'other'),
                'project': project_cache.get(project_name) if project_name else None,
                'project_name': project_name,
                'amount': float(amount),
                'date': date,
                'supplier': str(payee_val).strip() if payee_val else '',
                'status': status_map.get(str(status_val).strip(), 'draft') if status_val else 'draft',
                'description': str(desc_val).strip() if desc_val else '',
            })
            result.success += 1
        except Exception as e:
            result.add_error(i, f'行{i} 解析异常：{str(e)}')

    return result


# ─── 工资导入 ────────────────────────────────────────────────────────

def import_wage(file_obj, company_id=None):
    """
    导入工资记录。
    必需列：员工姓名, 年月 (如 2026-04), 公司
    可选列：基本工资, 加班费, 奖金, 社保个人, 公积金个人, 个税, 其他扣款, 实发工资
    """
    result = ImportResult()
    headers, all_rows = parse_uploaded_file(file_obj)

    col_name    = match_header(headers, '姓名', '员工姓名', 'name', '员工')
    col_yyyymm  = match_header(headers, '年月', 'year_month', '月份', '工资月份')
    col_company = match_header(headers, '公司', 'company', '所属公司')
    col_basic   = match_header(headers, '基本工资', 'basic_wage', '岗位工资')
    col_overtime= match_header(headers, '加班费', 'overtime_wage', '加班工资')
    col_bonus   = match_header(headers, '奖金', 'bonus', '绩效奖')
    col_soc_ins = match_header(headers, '社保', '养老保险', '个人社保', 'social_insurance')
    col_housing = match_header(headers, '公积金', '住房公租金', 'housing_fund', 'housing')
    col_tax     = match_header(headers, '个税', '个人所得税', 'personal_income_tax', 'tax')
    col_other   = match_header(headers, '其他扣款', 'other_deductions', '扣款')
    col_net     = match_header(headers, '实发', '实发工资', 'net_salary', '税后工资')

    if col_name == -1:
        result.add_error(0, '未找到"姓名"列，请确保 Excel 包含员工姓名列')
        return result
    if col_yyyymm == -1:
        result.add_error(0, '未找到"年月"列，请确保 Excel 包含年月列（如 2026-04）')
        return result
    if col_company == -1:
        result.add_error(0, '未找到"公司"列')
        return result

    from apps.finance.models import Employee, Company, WageRecord, EmployeeCompany

    # 公司名简称→全称映射（解决 Excel 模板用简称但数据库用全称的问题）
    COMPANY_ALIAS_MAP = {
        '百川': '深圳市百川软件科技发展有限公司',
        '百川实业': '深圳市百川软件科技发展有限公司',
        '百川软件': '深圳市百川软件科技发展有限公司',
        '龙晟': '深圳市龙晟科技贸易有限公司',
        '龙晟建筑': '深圳市龙晟科技贸易有限公司',
        '龙晟科技': '深圳市龙晟科技贸易有限公司',
        '绿聚能': '深圳市绿聚能科技有限公司',
        '绿聚': '深圳市绿聚能科技有限公司',
    }

    def resolve_company(name):
        if not name:
            return None
        s = name.strip()
        if s in company_cache:
            return company_cache[s]
        if s in COMPANY_ALIAS_MAP:
            return company_cache.get(COMPANY_ALIAS_MAP[s])
        # 模糊匹配：取包含
        for k, v in COMPANY_ALIAS_MAP.items():
            if k in s or s in k:
                return company_cache.get(v)
        for cname, cid in company_cache.items():
            if cname and s in cname:
                return cid
        return None

    # 员工缓存（name -> employee obj）
    emp_cache = {}
    for emp in Employee.objects.all():
        emp_cache[emp.name] = emp

    # 公司缓存（name -> id）
    company_cache = {c.name: c.id for c in Company.objects.all()}
    # 公司 id -> name 反查
    company_id_to_name = {c.id: c.name for c in Company.objects.all()}

    # EmployeeCompany 缓存（(emp_name, company_name) -> ec obj）
    ec_cache = {}
    for ec in EmployeeCompany.objects.select_related('employee', 'company').all():
        ec_cache[(ec.employee.name, ec.company.name)] = ec

    def parse_yyyymm(val):
        if not val:
            return None
        s = str(val).strip()
        # 2026-04 / 2026/04 / 202604 / 2026年04月
        for fmt in ('%Y-%m', '%Y/%m', '%Y%m', '%Y年%m'):
            try:
                return datetime.datetime.strptime(s, fmt).strftime('%Y-%m')
            except ValueError:
                pass
        return None

    for i, row in enumerate(all_rows, start=2):
        try:
            name = row[col_name] if col_name >= 0 and col_name < len(row) else None
            name = name.strip() if name else None
            if not name:
                result.add_error(i, f'行{i}：员工姓名不能为空')
                continue

            # 尝试解析年月（支持：'2026-04' / '2026,04' / 整数月份）
            raw_yyyymm = row[col_yyyymm] if col_yyyymm >= 0 and col_yyyymm < len(row) else None
            yyyymm = parse_yyyymm(raw_yyyymm)
            # 如果年月是分离的两列：'年份'=2026 在左，'月份'=12 在右
            if not yyyymm and raw_yyyymm is not None:
                try:
                    mo = int(float(str(raw_yyyymm).strip()))
                    if 1 <= mo <= 12:
                        # col_yyyymm 是月份列（右侧），年份列在左侧一格
                        yr_col = col_yyyymm - 1
                        if 0 <= yr_col < len(row) and row[yr_col] is not None:
                            yr = int(float(str(row[yr_col]).strip()))
                            if 2000 <= yr <= 2100:
                                yyyymm = f'{yr}-{mo:02d}'
                except (ValueError, TypeError):
                    pass
            if not yyyymm:
                result.add_error(i, f'行{i}：年月格式无效（请使用 2026-04 格式）')
                continue

            company_name = row[col_company] if col_company >= 0 and col_company < len(row) else None
            company_name = company_name.strip() if company_name else None
            company_id_val = resolve_company(company_name) if company_name else None

            # 查找员工
            employee = emp_cache.get(name)
            # 通过 company_id 查到公司全称，再用全称查 EmployeeCompany
            company_full_name = company_id_to_name.get(company_id_val)
            employee_company = ec_cache.get((name, company_full_name)) if name and company_full_name else None

            def get_num(col):
                if col < 0 or col >= len(row): return None
                return parse_number(row[col])

            result.rows.append({
                'employee': employee.id if employee else None,
                'employee_company': employee_company.id if employee_company else None,
                'employee_name': name,
                'company': company_id_val,
                'company_name': company_name,
                'year_month': yyyymm,
                'basic_wage': get_num(col_basic),
                'overtime_wage': get_num(col_overtime),
                'bonus': get_num(col_bonus),
                'social_insurance_employee': get_num(col_soc_ins),
                'housing_fund_employee': get_num(col_housing),
                'personal_income_tax': get_num(col_tax),
                'other_deductions': get_num(col_other),
                'net_salary': get_num(col_net),
            })
            result.success += 1
        except Exception as e:
            result.add_error(i, f'行{i} 解析异常：{str(e)}')

    return result

# ─── 发票导入 ─────────────────────────────────────────────────────────

def import_invoice(file_obj, invoice_type, company_id=None, operator=None):
    """
    导入数电发票记录（收到或开出的发票）。

    invoice_type: 'income' = 收到的发票（我方是购方，从供应商取得）
                  'expense' = 开出的发票（我方是销方，开给客户）

    Excel 格式：标准数电发票「发票基础信息」Sheet
    """
    result = ImportResult()
    data = file_obj.read() if hasattr(file_obj, 'read') else file_obj
    wb = load_workbook(io.BytesIO(data), data_only=True)

    # 优先找「发票基础信息」sheet
    ws = None
    for sheet_name in wb.sheetnames:
        if '发票基础信息' in sheet_name:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    # 找到表头行（含「数电发票号码」关键字）
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell is not None for cell in row) and '数电发票号码' in str(row):
            header_row_idx = i
            break
    if header_row_idx is None:
        result.add_error(0, '未找到「数电发票号码」表头，请确认文件格式')
        return result

    header_row = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    headers = [str(h).strip() if h is not None else '' for h in header_row]

    def col(*kw):
        return match_header(headers, *kw)

    c_invoice_no       = col('数电发票号码')
    c_seller_name      = col('销方名称')
    c_seller_tax       = col('销方识别号')
    c_buyer_name       = col('购买方名称')
    c_buyer_tax        = col('购方识别号')
    c_date             = col('开票日期')
    c_amount           = col('金额')
    c_tax              = col('税额')
    c_total            = col('价税合计')
    c_status           = col('发票状态')
    c_invoice_type_col = col('发票票种')
    c_issuer           = col('开票人')
    c_remark           = col('备注')

    def get_val(row, ci):
        if ci < 0 or ci >= len(row):
            return None
        v = row[ci]
        if v is None:
            return None
        if isinstance(v, float) and v != int(v):
            return round(v, 2)
        return v

    def parse_date(v):
        if v is None:
            return None
        if hasattr(v, 'strftime'):
            try:
                return v.date()
            except Exception:
                return v
        s = str(v).strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y%m%d'):
            try:
                return datetime.datetime.strptime(s[:10], fmt[:len(s)]).date()
            except ValueError:
                pass
        return None

    def map_invoice_type(v):
        if not v:
            return 'normal'
        s = str(v)
        if '专用' in s:
            return 'special'
        return 'normal'

    def map_status(status_val, inv_type):
        if not status_val:
            return 'paid' if inv_type == 'income' else 'pending'
        s = str(status_val)
        if '正常' in s:
            return 'paid' if inv_type == 'income' else 'pending'
        if '红冲' in s or '作废' in s:
            return 'cancelled'
        return 'paid' if inv_type == 'income' else 'pending'

    def resolve_company_by_tax(tax_id, default=None):
        from apps.core.models import Company
        if not tax_id:
            return default
        company = Company.objects.filter(tax_id=str(tax_id).strip()).first()
        return company.id if company else default

    # 聚合：同发票号多行明细合并为一条
    invoice_map = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue
        inv_no = get_val(row, c_invoice_no)
        if not inv_no:
            continue
        inv_no = str(inv_no).strip()
        if inv_no not in invoice_map:
            invoice_map[inv_no] = {
                'rows': [row],
                'seller_name': get_val(row, c_seller_name),
                'seller_tax':  get_val(row, c_seller_tax),
                'buyer_name':  get_val(row, c_buyer_name),
                'buyer_tax':   get_val(row, c_buyer_tax),
                'date':        parse_date(get_val(row, c_date)),
                '票种':        get_val(row, c_invoice_type_col),
                'issuer':      get_val(row, c_issuer),
                'remark':      get_val(row, c_remark),
                'status_raw':  get_val(row, c_status),
            }
        else:
            invoice_map[inv_no]['rows'].append(row)

    if invoice_type == 'income':
        counterparty_field    = 'seller_name'
        counterparty_tax_field = 'seller_tax'
    else:
        counterparty_field    = 'buyer_name'
        counterparty_tax_field = 'buyer_tax'

    from apps.finance.models import Invoice

    for inv_no, info in invoice_map.items():
        first_row = info['rows'][0]
        seq = first_row[0]

        try:
            # 判断正负（红冲票金额为负）
            is_negative = any(
                get_val(r, c_amount) is not None and float(get_val(r, c_amount)) < 0
                for r in info['rows']
            )

            # 汇总金额（取绝对值）
            total_amount = 0.0
            total_tax = 0.0
            for r in info['rows']:
                amt = get_val(r, c_amount)
                tax = get_val(r, c_tax)
                if amt is not None:
                    total_amount += abs(float(amt))
                if tax is not None:
                    total_tax += abs(float(tax))

            # 若金额列为0但价税合计有值，用价税合计反推
            if total_amount == 0:
                raw_total = get_val(first_row, c_total)
                if raw_total:
                    total_amount = abs(float(raw_total))
                    total_tax = 0.0

            counterparty       = str(info[counterparty_field] or '').strip()
            counterparty_tax_id = str(info[counterparty_tax_field] or '').strip()
            issue_date = info['date']

            # 公司匹配
            if invoice_type == 'income':
                resolved_company_id = resolve_company_by_tax(
                    info.get('buyer_tax'), default=company_id)
            else:
                resolved_company_id = resolve_company_by_tax(
                    info.get('seller_tax'), default=company_id)

            inv_status     = map_status(info.get('status_raw'), invoice_type)
            inv_invoice_type = map_invoice_type(info.get('票种'))

            # 备注拼接开票人
            issuer = str(info.get('issuer') or '').strip()
            remark_base = str(info.get('remark') or '').strip()
            remarks = (remark_base + (' | 开票人:' + issuer) if issuer else remark_base)

            # 跳过已存在的发票号
            if Invoice.objects.filter(invoice_no=inv_no).exists():
                result.add_error(seq, f'发票号 {inv_no} 已存在，跳过')
                continue

            result.rows.append({
                'invoice_no':          inv_no,
                'type':                invoice_type,
                'invoice_type':        inv_invoice_type,
                'amount':              round(total_amount, 2),
                'tax_amount':          round(total_tax, 2),
                'counterparty':        counterparty,
                'counterparty_tax_id': counterparty_tax_id,
                'issue_date':          issue_date,
                'status':              inv_status,
                'is_credited':         (inv_status != 'cancelled'),
                'company':             resolved_company_id or company_id,
                'remarks':             remarks,
            })
            result.success += 1

        except Exception as e:
            result.add_error(seq, f'发票 {inv_no} 解析异常：{str(e)}')

    return result
    
