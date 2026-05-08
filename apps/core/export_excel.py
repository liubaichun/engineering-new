"""
Excel 导出工具 — 使用 openpyxl 生成专业格式的 .xlsx 文件
支持：自动列宽、多表、样式（标题/金额/日期）、流式下载（大文件不爆内存）
"""
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter


# ─── 颜色常量 ──────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="4F46E5")   # 主色：靛蓝
SUBHDR_FILL   = PatternFill("solid", fgColor="EEF2FF")   # 淡紫
ALT_ROW_FILL  = PatternFill("solid", fgColor="F8FAFC")   # 交替行
WARN_FILL     = PatternFill("solid", fgColor="FEF3C7")   # 警告行
OK_FILL       = PatternFill("solid", fgColor="DCFCE7")   # 绿色
DANGER_FILL   = PatternFill("solid", fgColor="FEE2E2")   # 红色

WHITE_FONT    = Font(name='Microsoft YaHei', bold=True, color="FFFFFF", size=11)
DARK_FONT     = Font(name='Microsoft YaHei', bold=True, color="1E293B", size=10)
NORMAL_FONT   = Font(name='Microsoft YaHei', size=10)
SMALL_FONT    = Font(name='Microsoft YaHei', size=9, color="64748B")

CENTER        = Alignment(horizontal='center', vertical='center')
LEFT          = Alignment(horizontal='left', vertical='center')
RIGHT         = Alignment(horizontal='right', vertical='center')

THIN_BORDER   = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)


def _border():
    return THIN_BORDER


def _col_width(col_idx, header, sample_rows):
    """根据表头和示例数据自动计算列宽"""
    max_len = len(str(header)) if header else 8
    for row in sample_rows[:10]:
        val = row[col_idx] if col_idx < len(row) else ''
        max_len = max(max_len, len(str(val)))
    return min(max_len + 4, 60)


# ─── 核心导出函数 ──────────────────────────────────────────────────────

def export_to_xlsx(sheets: list, filename: str = None) -> io.BytesIO:
    """
    多表导出器。
    sheets: [
        {
            'title': 'Sheet名称',
            'headers': ['列1', '列2', ...],
            'rows': [[val1, val2, ...], ...],
            'column_types': ['text', 'money', 'date', 'percent', 'text'],  # 可选
            'freeze': 'A2',   # 可选，冻结行/列
        }
    ]
    返回 BytesIO，可直接作为 Django HttpResponse
    """
    wb = Workbook()
    wb.iso_dates = True

    for sheet_idx, sheet_conf in enumerate(sheets):
        if sheet_idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()

        title  = sheet_conf.get('title', f'Sheet{sheet_idx+1}')
        headers = sheet_conf.get('headers', [])
        rows    = sheet_conf.get('rows', [])
        col_types = sheet_conf.get('column_types', ['text'] * len(headers))
        freeze   = sheet_conf.get('freeze', 'A2')

        ws.title = title[:31]

        # 标题行（跨所有列）
        if title:
            ws.row_dimensions[1].height = 32
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font   = Font(name='Microsoft YaHei', bold=True, color="FFFFFF", size=13)
            title_cell.fill   = PatternFill("solid", fgColor="4F46E5")
            title_cell.alignment = CENTER

        # 表头行
        hdr_row = 2
        ws.row_dimensions[hdr_row].height = 22
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=hdr_row, column=col_idx, value=header)
            cell.font      = WHITE_FONT
            cell.fill      = HEADER_FILL if title else SUBHDR_FILL
            cell.alignment = CENTER
            cell.border    = _border()

        # 数据行
        for row_idx, row_data in enumerate(rows, start=hdr_row + 1):
            fill = ALT_ROW_FILL if (row_idx - hdr_row) % 2 == 0 else None
            ws.row_dimensions[row_idx].height = 18
            for col_idx, (raw_val, col_type) in enumerate(zip(row_data, col_types), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_format_value(raw_val, col_type))
                cell.border = _border()
                if fill:
                    cell.fill = fill

                # 样式
                if col_type == 'money':
                    cell.alignment = RIGHT
                    cell.number_format = '#,##0.00'
                    cell.font = Font(name='Microsoft YaHei', size=10)
                elif col_type == 'date':
                    cell.alignment = CENTER
                    cell.font = NORMAL_FONT
                elif col_type == 'percent':
                    cell.alignment = CENTER
                    cell.number_format = '0.0%'
                    cell.font = NORMAL_FONT
                elif col_type == 'status':
                    cell.alignment = CENTER
                    cell.font = Font(name='Microsoft YaHei', size=10,
                                     color=_status_color(str(raw_val)))
                elif col_type == 'bool':
                    cell.alignment = CENTER
                    cell.font = Font(name='Microsoft YaHei', size=10,
                                     color="166534" if raw_val else "991b1b")
                else:
                    cell.alignment = LEFT
                    cell.font = NORMAL_FONT

        # 自动列宽
        sample = rows[:20]
        for col_idx, header in enumerate(headers, start=1):
            width = _col_width(col_idx - 1, header, sample)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 冻结窗格
        if freeze:
            ws.freeze_panes = freeze

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _format_value(val, col_type):
    """将 Python 值格式化为 Excel 单元格的合理显示"""
    if val is None or val == '':
        return ''
    if col_type == 'money':
        try:
            return float(val)
        except (ValueError, TypeError):
            return val
    if col_type == 'date':
        if isinstance(val, datetime.datetime):
            if val.tzinfo is not None:
                val = val.replace(tzinfo=None)
            return val
        if isinstance(val, datetime.date):
            return datetime.datetime(val.year, val.month, val.day)
        # 字符串日期
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S',
                    '%Y-%m-%dT%H:%M:%S', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.datetime.strptime(str(val)[:19], fmt)
            except ValueError:
                continue
        return val
    if col_type == 'percent':
        try:
            v = float(val)
            return v / 100 if v > 1 else v
        except (ValueError, TypeError):
            return val
    if col_type == 'bool':
        return '是' if val in (True, 'true', 'True', 1, '1') else '否'
    if col_type == 'status':
        return str(val)
    return val


def _status_color(val):
    """根据状态值返回适合的字体颜色"""
    table = {
        '已确认': '166534', 'confirmed': '166534', 'approved': '166534',
        '已完成': '166534', 'completed': '166534', 'paid': '166534',
        '已支付': '166534',
        '待确认': '92400E', 'pending': '92400E', 'draft': '92400E',
        '待审核': '92400E', 'submitted': '92400E',
        '已拒绝': '991b1b', 'rejected': '991b1b', 'failed': '991b1b',
        '已作废': '991b1b', 'cancelled': '991b1b',
    }
    return table.get(val, '334155')


# ─── 专用导出函数 ────────────────────────────────────────────────────────

def export_wage_records(records, company_name=''):
    """导出工资单"""
    title = f'工资单 {company_name} {datetime.datetime.now().strftime("%Y年%m月")}'
    headers = ['序号', '姓名', '身份证', '公司', '部门', '岗位', '基本工资',
               '岗位工资', '加班费', '奖金', '社保扣款', '公积金扣款',
               '个税', '实发工资', '状态']
    rows = []
    for i, r in enumerate(records, 1):
        rows.append([
            i,
            r.employee_company.employee.name if r.employee_company and r.employee_company.employee else (r.employee.name if r.employee else r.employee_name or ''),
            r.employee_company.employee.id_card if r.employee_company and r.employee_company.employee else (r.employee.id_card if r.employee else ''),
            r.company.name if r.company else '',
            r.department or '',
            r.position or '',
            r.base_salary or 0, r.position_salary or 0,
            r.overtime_pay or 0, r.bonus or 0,
            r.social_insurance or 0, r.housing_fund or 0,
            r.tax or 0, r.net_salary or 0,
            r.get_status_display() if hasattr(r,'get_status_display') else '',
        ])
    col_types = ['text','text','text','text','text','text','money','money','money','money','money','money','money','money','status']
    wb_sheets = [{'title': title, 'headers': headers, 'rows': rows,
                  'column_types': col_types, 'freeze': 'A3'}]
    return export_to_xlsx(wb_sheets, filename=f'工资单_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_income_records(records):
    """导出收入记录"""
    title = f'收入记录 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = [
        'ID', '公司', '交易时间', '来源', '对手账号', '对手开户行',
        '金额(元)', '余额', '日期', '客户', '关联项目',
        '状态', '审批状态', '操作人', '备注', '创建时间'
    ]
    rows = [[
        r.id,
        r.company.name if r.company else '',
        r.transaction_time.strftime('%H:%M:%S') if r.transaction_time else '',
        r.source or '',
        r.counterparty_account or '',
        r.counterparty_bank or '',
        r.amount,
        f'{r.balance}' if r.balance else '',
        r.date or '',
        r.customer or '',
        r.project.name if r.project else '',
        r.get_status_display() if hasattr(r, 'get_status_display') else r.status or '',
        r.approval_flow.get_status_display() if r.approval_flow else '',
        r.operator.username if r.operator else '',
        r.description or '',
        r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
    ] for r in records]
    col_types = ['text', 'text', 'time', 'text', 'text', 'text', 'money', 'money', 'date', 'text', 'text', 'status', 'status', 'text', 'text', 'text']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'收入记录_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_expense_records(records):
    """导出支出记录"""
    title = f'支出记录 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = [
        'ID', '公司', '交易时间', '支出类型', '对手账号', '对手开户行',
        '金额(元)', '余额', '日期', '供应商', '关联项目',
        '状态', '审批状态', '操作人', '备注', '创建时间'
    ]
    rows = [[
        r.id,
        r.company.name if r.company else '',
        r.transaction_time.strftime('%H:%M:%S') if r.transaction_time else '',
        r.get_expense_type_display() if hasattr(r, 'get_expense_type_display') else '',
        r.counterparty_account or '',
        r.counterparty_bank or '',
        r.amount,
        f'{r.balance}' if r.balance else '',
        r.date or '',
        r.supplier or '',
        r.project.name if r.project else '',
        r.get_status_display() if hasattr(r, 'get_status_display') else r.status or '',
        r.approval_flow.get_status_display() if r.approval_flow else '',
        r.operator.username if r.operator else '',
        r.description or '',
        r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
    ] for r in records]
    col_types = ['text', 'text', 'time', 'text', 'text', 'text', 'money', 'money', 'date', 'text', 'text', 'status', 'status', 'text', 'text', 'text']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'支出记录_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_invoices(records):
    """导出发票"""
    title = f'发票管理 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '发票号码', '公司', '项目', '开票日期', '到期日期', '金额(元)', '税率',
               '税额(元)', '价税合计', '发票类型', '是否认证', '状态', '备注', '创建时间']
    rows = [[
        r.id, r.invoice_no or '', r.company.name if r.company else '',
        r.project.name if r.project else '',
        r.issue_date or '', r.due_date or '',
        r.amount or 0,
        f"{float(r.tax_rate)*100:.0f}%" if r.tax_rate else '',
        r.tax_amount or 0, (r.amount or 0) + (r.tax_amount or 0),
        r.get_invoice_type_display() if hasattr(r,'get_invoice_type_display') else '',
        '是' if r.is_credited else '否',
        r.get_status_display() if hasattr(r,'get_status_display') else r.status or '',
        r.remarks or '',
        r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
    ] for r in records]
    col_types = ['text','text','text','text','date','date','money','percent','money','money','text','bool','status','text','text']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'发票_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_employees(records):
    """导出员工"""
    title = f'员工花名册 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '姓名', '手机', '邮箱', '身份证', '入职日期',
               '部门', '岗位', '公司', '状态']
    rows = [[
        r.id, r.name,
        r.phone or '', r.email or '', r.id_card or '',
        r.hire_date or '', r.department or '', r.position or '',
        r.company.name if r.company else '',
        r.get_status_display() if hasattr(r,'get_status_display') else '',
    ] for r in records]
    col_types = ['text','text','text','text','text','date','text','text','text','status']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'员工_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_companies(records):
    """导出公司"""
    title = f'公司管理 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '公司名称', '公司代码', '地址', '联系人', '联系电话',
               '开户银行', '银行账号', '状态', '创建时间']
    rows = [[
        r.id, r.name, r.code or '',
        r.address or '', r.contact_person or '', r.contact_phone or '',
        r.bank_name or '', r.bank_account or '',
        r.get_status_display() if hasattr(r,'get_status_display') else '',
        r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
    ] for r in records]
    col_types = ['text','text','text','text','text','text','text','text','status','date']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'公司_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_clients(records):
    """导出客户"""
    title = f'客户列表 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '客户名称', '联系人', '电话', '地址', '备注', '状态', '创建时间']
    rows = [[
        r.id, r.name, r.contact_person or '',
        r.contact_phone or '', r.address or '',
        r.remark or '',
        r.is_active if hasattr(r,'is_active') else True,
        r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
    ] for r in records]
    col_types = ['text','text','text','text','text','text','bool','date']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'客户_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_contracts(records):
    """导出合同"""
    title = f'合同管理 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '合同编号', '客户', '合同名称', '金额(元)', '签署日期',
               '到期日期', '状态', '备注', '创建时间']
    rows = []
    today = datetime.date.today()
    for r in records:
        end = r.expire_date
        days_left = (end - today).days if end else None
        rows.append([
            r.id, r.contract_no or '',
            r.client.name if r.client else '',
            r.name, r.amount or 0,
            r.sign_date or '', r.expire_date or '',
            r.get_status_display() if hasattr(r,'get_status_display') else '',
            r.remark or '',
            r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
        ])
    col_types = ['text','text','text','text','money','date','date','status','text','date']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'合同_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_projects(records):
    """导出项目"""
    title = f'项目列表 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '项目名称', '项目代码', '负责人', '公司',
               '开始日期', '结束日期', '预算(元)',
               '项目进度(%)', '状态', '创建时间']
    rows = [[
        r.id, r.name, r.code or '',
        r.owner.username if r.owner else '',
        r.company.name if r.company else '',
        r.start_date or '', r.end_date or '',
        r.budget or 0,
        round(float(r.progress or 0), 1),
        r.get_status_display() if hasattr(r,'get_status_display') else '',
        r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
    ] for r in records]
    col_types = ['text','text','text','text','text','date','date','money','percent','status','date']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'项目_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_equipment(records):
    """导出设备"""
    title = f'设备管理 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '设备名称', '设备编码', '分类', '管理方式', '序列号',
               '批次号', '单位', '购入日期', '保修截止', '购入价格(元)',
               '存放地点', '状态', '备注']
    rows = [[
        r.id, r.name, r.code or '',
        r.get_category_display() if hasattr(r,'get_category_display') else '',
        r.get_management_type_display() if hasattr(r,'get_management_type_display') else '',
        r.serial_number or '', r.batch_number or '',
        r.unit or '', r.purchase_date or '', r.warranty_end or '',
        r.purchase_price or 0,
        r.location or '',
        r.get_status_display() if hasattr(r,'get_status_display') else '',
        r.remarks or '',
    ] for r in records]
    col_types = ['text','text','text','text','text','text','text','text','date','date','money','text','status','text']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'设备_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_suppliers(records):
    """导出供应商"""
    title = f'供应商列表 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '编码', '名称', '联系人', '联系电话', '邮箱',
               '代理品牌', '合作状态', '地址', '备注', '创建时间']
    rows = [[
        r.id, r.code or '', r.name,
        r.contact_person or '', r.contact_phone or '', r.contact_email or '',
        r.brands or '', r.get_status_display() if hasattr(r,'get_status_display') else '',
        r.address or '', r.remark or '',
        r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
    ] for r in records]
    col_types = ['text','text','text','text','text','text','text','status','text','text','date']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'供应商_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


def export_audit_logs(logs):
    """导出审计日志"""
    title = f'操作审计日志 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '操作用户', '操作', '应用', '模型', '对象ID',
               '对象摘要', 'IP地址', '审批流ID', '操作时间']
    rows = [[
        r.id, r.username,
        r.get_action_display() if hasattr(r,'get_action_display') else r.action,
        r.app_label, r.model_name, r.object_id,
        r.object_repr or '', r.ip_address or '',
        r.approval_flow_id or '',
        r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
    ] for r in logs]
    col_types = ['text','text','status','text','text','text','text','text','text','text']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'审计日志_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')


# ─── 通用 HTTP Response ─────────────────────────────────────────────────

def make_export_response(buf: io.BytesIO, filename: str) -> 'HttpResponse':
    """将 BytesIO 包装为 Django HttpResponse"""
    from django.http import HttpResponse
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return response


def export_materials(records):
    """导出物料"""
    title = f'物料管理 {datetime.datetime.now().strftime("%Y年%m月%d日")}'
    headers = ['ID', '物料编码', '物料名称', '规格型号', '分类', '单位',
               '库存', '预警阈值', '单价(元)', '供应商', '备注', '创建时间']
    rows = [[
        r.id, r.code or '', r.name or '',
        r.spec or '',
        r.category or '',
        r.unit or '',
        r.stock or 0,
        r.alert_threshold or 0,
        r.unit_price or 0,
        r.supplier.name if r.supplier else '',
        r.remark or '',
        r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
    ] for r in records]
    col_types = ['text','text','text','text','text','text','number','number','money','text','text','date']
    return export_to_xlsx([{'title': title, 'headers': headers, 'rows': rows,
                             'column_types': col_types, 'freeze': 'A3'}],
                           filename=f'物料_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')
