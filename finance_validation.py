#!/usr/bin/env python3
"""
财务模块深度验证：工资+社保+税务+导入导出
重点验证：
1. 工资API全部200
2. net公式13条验证
3. 2026+2027 tax>0
4. 11个导出API全部200
5. 空文件导入400拒绝
"""
import io, sys, os, time, json
from datetime import date
sys.path.insert(0, '/root/engineering-new')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

import django; django.setup()

import requests
import openpyxl
from django.db import connection, reset_queries

BASE = 'http://127.0.0.1:8001'
SESSION = requests.Session()

resp = SESSION.post(f'{BASE}/api/core/auth/login/', json={'username': 'admin', 'password': 'admin123'})
if resp.status_code != 200:
    print(f'❌ 登录失败: {resp.status_code}'); sys.exit(1)
print(f'✅ 登录成功')

PASS = 0; FAIL = 0

def rec(name, ok, detail=''):
    global PASS, FAIL
    icon = '✅' if ok else '❌'
    if ok: PASS += 1
    else: FAIL += 1
    print(f'{icon} {name}' + (f' — {detail}' if detail else ''))

def get(url):
    r = SESSION.get(url); return r.status_code, r.content

def post(url, **kw):
    r = SESSION.post(url, **kw); return r.status_code, r.content

def is_xlsx(content_bytes):
    if not content_bytes or len(content_bytes) < 4: return False
    if content_bytes[:2] == b'PK':
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes)); return bool(wb.sheetnames)
        except: return False
    return False

def excel_bytes(rows, headers):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(headers)
    for r in rows: ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ============================================================
# 第一部分：工资API全部200 (wage endpoints)
# ============================================================
print('\n' + '='*60)
print('【第一部分】工资API全部200')
print('='*60)

wage_apis = [
    ('GET 工资列表',   '/api/finance/wages/'),
    ('GET 工资明细',   '/api/finance/wages/1/'),
    ('GET 工资汇总',   '/api/finance/wages/summary/'),
    ('GET 工资仪表盘', '/api/finance/wages/dashboard/'),
    ('GET 工资导出',   '/api/finance/wages/export/'),
]

for name, path in wage_apis:
    sc, content = get(BASE + path)
    xlsx = is_xlsx(content) if 'export' in path else True
    rec(name, sc == 200, f'HTTP {sc} {len(content)}B')

# POST actions
sc, _ = post(f'{BASE}/api/finance/wages/', json={})
rec('POST 工资(空数据)', sc in (200,201,400), f'HTTP {sc}')

# ============================================================
# 第二部分：net公式13条验证
# WageRecord.calculate_gross_and_tax() 逐行验证
# ============================================================
print('\n' + '='*60)
print('【第二部分】net公式13条验证')
print('='*60)

from apps.finance.models import WageRecord, Company, Employee, EmployeeCompany
from decimal import Decimal

# 获取测试数据
co = Company.objects.first()
ec = EmployeeCompany.objects.select_related('employee', 'company').first()
emp = ec.employee if ec else None
if not co:
    print("❌ 没有公司数据"); sys.exit(1)
if not emp:
    print("❌ 没有员工数据"); sys.exit(1)

def calc_tax(taxable):
    """累计预扣法税率计算（复制 WageRecord.calculate_gross_and_tax 逻辑）"""
    if taxable <= 0:
        return 0.0
    thresholds = [0, 3000, 12000, 25000, 35000, 55000, 80000]
    rates = [3, 10, 20, 25, 30, 35, 45]
    quick_deductions = [0, 210, 1410, 2660, 4410, 7160, 15160]
    for i in range(len(thresholds) - 1):
        if taxable <= thresholds[i + 1]:
            return max(taxable * rates[i] / 100 - quick_deductions[i], 0)
    return max(taxable * 45 / 100 - 15160, 0)

def gross_formula(base=0, position=0, overtime=0, bonus=0, commission=0,
                  meal=0, transport=0, comm=0, other=0):
    return (base + position + overtime + bonus + commission +
            meal + transport + comm + other)

def total_ded_formula(soc_ins=0, housing=0, leave=0, sick=0,
                      loan=0, other_loan=0, other=0):
    return soc_ins + housing + leave + sick + loan + other_loan + other

# 测试用例：13条公式验证
test_cases = []

# 场景A: 2026年1月 基准测试（无累计）
gross_a = gross_formula(base=8000, position=2000, bonus=1000)
total_ded_a = total_ded_formula(soc_ins=561, housing=225)
taxable_a = gross_a - total_ded_a - 5000
tax_a = calc_tax(taxable_a)
net_a = gross_a - total_ded_a - tax_a

# 清理2026年1月的旧测试数据
WageRecord.objects.filter(year=2026, month=1, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
wr_a = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=1,
    base_salary=Decimal('8000'), position_salary=Decimal('2000'),
    bonus=Decimal('1000'), social_insurance=Decimal('561'),
    housing_fund=Decimal('225'),
    status='draft'
)
test_cases.append(('2026-01 基准', wr_a, gross_a, total_ded_a, tax_a, net_a))

# 场景B: 2026年2月 累进验证（与1月累计）
gross_b = gross_formula(base=8000, position=2000, bonus=1200)
total_ded_b = total_ded_formula(soc_ins=561, housing=225)
taxable_b = (gross_a + gross_b) - (total_ded_a + total_ded_b) - 5000 * 2
tax_b = calc_tax(taxable_b) - calc_tax(taxable_a)  # 当月=累计-上月
net_b = gross_b - total_ded_b - tax_b

WageRecord.objects.filter(year=2026, month=2, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
wr_b = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=2,
    base_salary=Decimal('8000'), position_salary=Decimal('2000'),
    bonus=Decimal('1200'), social_insurance=Decimal('561'),
    housing_fund=Decimal('225'),
    status='draft'
)
test_cases.append(('2026-02 累进', wr_b, gross_b, total_ded_b, tax_b, net_b))

# 场景C: 2026年3月 工资不变，社保公积金不变，tax 应不变
gross_c = gross_formula(base=8000, position=2000, bonus=1200)
total_ded_c = total_ded_formula(soc_ins=561, housing=225)
taxable_c = (gross_a + gross_b + gross_c) - (total_ded_a + total_ded_b + total_ded_c) - 5000 * 3
tax_c = calc_tax(taxable_c) - calc_tax(taxable_b)
net_c = gross_c - total_ded_c - tax_c

WageRecord.objects.filter(year=2026, month=3, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
wr_c = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=3,
    base_salary=Decimal('8000'), position_salary=Decimal('2000'),
    bonus=Decimal('1200'), social_insurance=Decimal('561'),
    housing_fund=Decimal('225'),
    status='draft'
)
test_cases.append(('2026-03 稳定', wr_c, gross_c, total_ded_c, tax_c, net_c))

# 场景D: 2027年1月 新年归零
gross_d = gross_formula(base=10000, position=3000, bonus=2000)
total_ded_d = total_ded_formula(soc_ins=800, housing=300)
taxable_d = gross_d - total_ded_d - 5000
tax_d = calc_tax(taxable_d)
net_d = gross_d - total_ded_d - tax_d

WageRecord.objects.filter(year=2027, month=1, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
wr_d = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2027, month=1,
    base_salary=Decimal('10000'), position_salary=Decimal('3000'),
    bonus=Decimal('2000'), social_insurance=Decimal('800'),
    housing_fund=Decimal('300'),
    status='draft'
)
test_cases.append(('2027-01 新年归零', wr_d, gross_d, total_ded_d, tax_d, net_d))

# 场景E: 2027年2月 高收入累进
gross_e = gross_formula(base=15000, position=5000, bonus=3000)
total_ded_e = total_ded_formula(soc_ins=800, housing=300)
taxable_e = (gross_d + gross_e) - (total_ded_d + total_ded_e) - 5000 * 2
tax_e = calc_tax(taxable_e) - calc_tax(taxable_d)
net_e = gross_e - total_ded_e - tax_e

WageRecord.objects.filter(year=2027, month=2, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
wr_e = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2027, month=2,
    base_salary=Decimal('15000'), position_salary=Decimal('5000'),
    bonus=Decimal('3000'), social_insurance=Decimal('800'),
    housing_fund=Decimal('300'),
    status='draft'
)
test_cases.append(('2027-02 高收入累进', wr_e, gross_e, total_ded_e, tax_e, net_e))

# 验证所有13个公式
for label, wr, exp_gross, exp_ded, exp_tax, exp_net in test_cases:
    wr.refresh_from_db()
    act_gross = float(wr.gross_salary or 0)
    act_ded = float(wr.total_deduction or 0)
    act_tax = float(wr.tax or 0)
    act_net = float(wr.net_salary or 0)

    g_ok = abs(act_gross - exp_gross) < 0.01
    d_ok = abs(act_ded - exp_ded) < 0.01
    t_ok = abs(act_tax - exp_tax) < 0.01
    n_ok = abs(act_net - exp_net) < 0.01

    rec(f'公式 {label} gross', g_ok,
        f'期望{exp_gross:.2f} 实际{act_gross:.2f}')
    rec(f'公式 {label} ded', d_ok,
        f'期望{exp_ded:.2f} 实际{act_ded:.2f}')
    rec(f'公式 {label} tax', t_ok,
        f'期望{exp_tax:.2f} 实际{act_tax:.2f}')
    rec(f'公式 {label} net', n_ok,
        f'期望{exp_net:.2f} 实际{act_net:.2f}')

# 公式10-13: 边界条件
# 场景F: 0社保公积金（最低档）
WageRecord.objects.filter(year=2026, month=4, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_f = gross_formula(base=5000, position=0)
total_ded_f = 0
taxable_f = gross_f - 0 - 5000
tax_f = calc_tax(taxable_f)
net_f = gross_f - total_ded_f - tax_f
wr_f = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=4, base_salary=Decimal('5000'),
    status='draft'
)
test_cases.append(('2026-04 0社保公积金', wr_f, gross_f, total_ded_f, tax_f, net_f))

# 场景G: 高额税（35000+应税收入）
WageRecord.objects.filter(year=2026, month=5, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_g = gross_formula(base=30000, position=10000, bonus=8000)
total_ded_g = total_ded_formula(soc_ins=2000, housing=1000)
taxable_g = gross_g - total_ded_g - 5000
tax_g = calc_tax(taxable_g)
net_g = gross_g - total_ded_g - tax_g
wr_g = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=5,
    base_salary=Decimal('30000'), position_salary=Decimal('10000'),
    bonus=Decimal('8000'), social_insurance=Decimal('2000'),
    housing_fund=Decimal('1000'),
    status='draft'
)
test_cases.append(('2026-05 高额税', wr_g, gross_g, total_ded_g, tax_g, net_g))

# 场景H: 所有补贴
WageRecord.objects.filter(year=2026, month=6, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_h = gross_formula(base=8000, position=2000, overtime=500, bonus=1000,
                         commission=500, meal=300, transport=200, comm=100, other=200)
total_ded_h = total_ded_formula(soc_ins=561, housing=225)
taxable_h = gross_h - total_ded_h - 5000
tax_h = calc_tax(taxable_h)
net_h = gross_h - total_ded_h - tax_h
wr_h = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=6,
    base_salary=Decimal('8000'), position_salary=Decimal('2000'),
    overtime_pay=Decimal('500'), bonus=Decimal('1000'),
    commission=Decimal('500'), meal_allowance=Decimal('300'),
    transport_allowance=Decimal('200'), communication_allowance=Decimal('100'),
    other_allowance=Decimal('200'),
    social_insurance=Decimal('561'), housing_fund=Decimal('225'),
    status='draft'
)
test_cases.append(('2026-06 全补贴', wr_h, gross_h, total_ded_h, tax_h, net_h))

# 场景I: 有其他扣款
WageRecord.objects.filter(year=2026, month=7, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_i = gross_formula(base=8000, position=2000, bonus=500)
total_ded_i = total_ded_formula(soc_ins=561, housing=225, leave=200, other=100)
taxable_i = gross_i - total_ded_i - 5000
tax_i = calc_tax(taxable_i)
net_i = gross_i - total_ded_i - tax_i
wr_i = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=7,
    base_salary=Decimal('8000'), position_salary=Decimal('2000'),
    bonus=Decimal('500'), social_insurance=Decimal('561'),
    housing_fund=Decimal('225'), leave_deduction=Decimal('200'),
    other_deductions=Decimal('100'),
    status='draft'
)
test_cases.append(('2026-07 其他扣款', wr_i, gross_i, total_ded_i, tax_i, net_i))

# 场景J: 应税收入为0（月薪=社保公积金）
WageRecord.objects.filter(year=2026, month=8, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_j = gross_formula(base=5561)  # gross = 5561
total_ded_j = 5561  # 社保+公积金恰好=gross
taxable_j = 0  # 5561-5561-5000 < 0 → 0
tax_j = 0.0
net_j = gross_j - total_ded_j - tax_j
wr_j = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2026, month=8, base_salary=Decimal('5561'),
    social_insurance=Decimal('5561'), housing_fund=Decimal('0'),
    status='draft'
)
test_cases.append(('2026-08 应税=0', wr_j, gross_j, total_ded_j, tax_j, net_j))

# 场景K: 2027 跨年累计验证
WageRecord.objects.filter(year=2027, month=3, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_k = gross_formula(base=10000, position=3000, bonus=1500)
total_ded_k = total_ded_formula(soc_ins=800, housing=300)
# 2027年1-3月累计
cum_gross_2027 = gross_d + gross_e + gross_k
cum_special_k = total_ded_d + total_ded_e + total_ded_k
taxable_k = cum_gross_2027 - cum_special_k - 5000 * 3
tax_k = calc_tax(taxable_k) - calc_tax(taxable_e)  # 3月当月=累计3月-累计2月
net_k = gross_k - total_ded_k - tax_k
wr_k = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2027, month=3,
    base_salary=Decimal('10000'), position_salary=Decimal('3000'),
    bonus=Decimal('1500'), social_insurance=Decimal('800'),
    housing_fund=Decimal('300'),
    status='draft'
)
test_cases.append(('2027-03 跨年累进', wr_k, gross_k, total_ded_k, tax_k, net_k))

# 场景L: 超高应税收入（超55000档）
WageRecord.objects.filter(year=2027, month=4, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_l = gross_formula(base=60000, position=20000, bonus=10000)
total_ded_l = total_ded_formula(soc_ins=3000, housing=1500)
cum_gross_l = gross_d + gross_e + gross_k + gross_l
cum_special_l = total_ded_d + total_ded_e + total_ded_k + total_ded_l
taxable_l = cum_gross_l - cum_special_l - 5000 * 4
tax_l = calc_tax(taxable_l) - calc_tax(taxable_k)
net_l = gross_l - total_ded_l - tax_l
wr_l = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2027, month=4,
    base_salary=Decimal('60000'), position_salary=Decimal('20000'),
    bonus=Decimal('10000'), social_insurance=Decimal('3000'),
    housing_fund=Decimal('1500'),
    status='draft'
)
test_cases.append(('2027-04 超高应税', wr_l, gross_l, total_ded_l, tax_l, net_l))

# 场景M: 只有基本工资=5000（个税=0）
WageRecord.objects.filter(year=2027, month=5, company=co, employee_name=emp.name,
                          status__in=['draft','pending']).delete()
gross_m = gross_formula(base=5000)
total_ded_m = 0
taxable_m = gross_m - 0 - 5000
tax_m = calc_tax(taxable_m)
net_m = gross_m - total_ded_m - tax_m
wr_m = WageRecord.objects.create(
    company=co, employee=emp, employee_name=emp.name,
    year=2027, month=5, base_salary=Decimal('5000'),
    status='draft'
)
test_cases.append(('2027-05 月薪=5000免税', wr_m, gross_m, total_ded_m, tax_m, net_m))

# 验证场景F-M（接续验证）
extra_scenarios = [
    ('2026-04 0社保公积金', wr_f, gross_f, total_ded_f, tax_f, net_f),
    ('2026-05 高额税', wr_g, gross_g, total_ded_g, tax_g, net_g),
    ('2026-06 全补贴', wr_h, gross_h, total_ded_h, tax_h, net_h),
    ('2026-07 其他扣款', wr_i, gross_i, total_ded_i, tax_i, net_i),
    ('2026-08 应税=0', wr_j, gross_j, total_ded_j, tax_j, net_j),
    ('2027-03 跨年累进', wr_k, gross_k, total_ded_k, tax_k, net_k),
    ('2027-04 超高应税', wr_l, gross_l, total_ded_l, tax_l, net_l),
    ('2027-05 月薪=5000免税', wr_m, gross_m, total_ded_m, tax_m, net_m),
]

for label, wr, exp_gross, exp_ded, exp_tax, exp_net in extra_scenarios:
    wr.refresh_from_db()
    act_gross = float(wr.gross_salary or 0)
    act_ded = float(wr.total_deduction or 0)
    act_tax = float(wr.tax or 0)
    act_net = float(wr.net_salary or 0)

    g_ok = abs(act_gross - exp_gross) < 0.01
    d_ok = abs(act_ded - exp_ded) < 0.01
    t_ok = abs(act_tax - exp_tax) < 0.01
    n_ok = abs(act_net - exp_net) < 0.01

    rec(f'公式 {label} gross', g_ok,
        f'期望{exp_gross:.2f} 实际{act_gross:.2f}')
    rec(f'公式 {label} ded', d_ok,
        f'期望{exp_ded:.2f} 实际{act_ded:.2f}')
    rec(f'公式 {label} tax', t_ok,
        f'期望{exp_tax:.2f} 实际{act_tax:.2f}')
    rec(f'公式 {label} net', n_ok,
        f'期望{exp_net:.2f} 实际{act_net:.2f}')

# ============================================================
# 第三部分：2026+2027 tax>0 验证
# ============================================================
print('\n' + '='*60)
print('【第三部分】2026+2027 tax>0 验证')
print('='*60)

# 找有tax记录的2026和2027工资记录
wages_2026 = WageRecord.objects.filter(year=2026, tax__gt=0).count()
wages_2027 = WageRecord.objects.filter(year=2027, tax__gt=0).count()
rec(f'2026年 tax>0 记录数', wages_2026 > 0, f'{wages_2026} 条')
rec(f'2027年 tax>0 记录数', wages_2027 > 0, f'{wages_2027} 条')

# 验证2026/2027每条 tax >= 0
for yr in [2026, 2027]:
    for wr in WageRecord.objects.filter(year=yr):
        wr.refresh_from_db()
        rec(f'{yr}-{wr.month:02d} tax>=0',
            float(wr.tax or 0) >= 0,
            f'tax={wr.tax}')

# ============================================================
# 第四部分：11个导出API全部200
# ============================================================
print('\n' + '='*60)
print('【第四部分】11个导出API全部200')
print('='*60)

exports = [
    ('公司导出',   '/api/finance/companies/export/'),
    ('收入导出',   '/api/finance/incomes/export/'),
    ('支出导出',   '/api/finance/expenses/export/'),
    ('工资导出',   '/api/finance/wages/export/'),
    ('发票导出',   '/api/finance/invoices/export/'),
    ('客户导出',   '/api/crm/clients/export/'),
    ('合同导出',   '/api/crm/contracts/export/'),
    ('供应商导出', '/api/crm/suppliers/export/'),
    ('项目导出',   '/api/tasks/projects/export/'),
    ('设备导出',   '/api/equipment/export/'),
    ('物料导出',   '/api/material/export/'),
]

for name, path in exports:
    sc, content = get(BASE + path)
    xlsx = is_xlsx(content)
    rec(name, sc == 200 and xlsx, f'HTTP {sc} Excel={xlsx} {len(content)}B')

# ============================================================
# 第五部分：空文件导入400拒绝
# ============================================================
print('\n' + '='*60)
print('【第五部分】空文件导入400拒绝')
print('='*60)

# 空Excel（只有表头无数据）
sc, content = post(f'{BASE}/api/finance/incomes/import_records/',
    files={'file': ('empty.xlsx',
        excel_bytes([], ['公司名称','金额','日期']),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
rec('空文件收入导入拒绝', sc == 400, f'HTTP {sc}')

sc, content = post(f'{BASE}/api/finance/expenses/import_records/',
    files={'file': ('empty.xlsx',
        excel_bytes([], ['公司名称','金额','日期']),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
rec('空文件支出导入拒绝', sc == 400, f'HTTP {sc}')

# 完全空文件（0字节）
sc, content = post(f'{BASE}/api/finance/wages/import_records/',
    files={'file': ('zero.xlsx', b'', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
rec('0字节文件工资导入拒绝', sc == 400, f'HTTP {sc}')

# 缺少必需列的空数据
sc, content = post(f'{BASE}/api/finance/incomes/import_records/',
    files={'file': ('bad.xlsx',
        excel_bytes([['','']], ['无效列1','无效列2']),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
rec('缺少必需列导入拒绝', sc == 400, f'HTTP {sc}')

# 无文件上传
sc = SESSION.post(f'{BASE}/api/finance/wages/import_records/').status_code
rec('无文件上传工资导入拒绝', sc == 400, f'HTTP {sc}')

# ============================================================
# 总结
# ============================================================
print('\n' + '='*60)
print(f'【验证完成】✅ {PASS} 项通过 | ❌ {FAIL} 项失败')
print('='*60)
sys.exit(0 if FAIL == 0 else 1)
