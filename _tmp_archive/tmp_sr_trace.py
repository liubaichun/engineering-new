"""
模拟导入一条缺失记录看实际发生了什么
"""
import os, sys
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.insert(0, '/root/engineering-new')
import django
django.setup()

from openpyxl import load_workbook
from apps.finance.models import SocialRecord, Employee, Company
from decimal import Decimal

excel_path = '/root/.hermes/profiles/hermes-b001/cache/documents/doc_e2de73598919_深圳市百川软件科技发展有限公司_社保费申报明细_20260527.xlsx2401.xlsx'
wb = load_workbook(excel_path, data_only=True)
ws = wb.active

# 看Sheet名
print(f'Sheet名: "{ws.title}"')

# 模拟公司解析逻辑
import re
sheet_name = ws.title
m = re.match(r'^(.+?)_社保费申报明细', sheet_name)
if m:
    print(f'匹配到公司名: {m.group(1)}')
else:
    print('未匹配到公司名（Sheet名格式不对）')
    # 会回退到第一个公司
    first = Company.objects.first()
    print(f'回退到第一个公司: {first.name} (id={first.id})')

# 查百川的Employee表有哪些人
bc = Company.objects.get(name__contains='百川')
emps = Employee.objects.filter(company=bc)
print()
print(f'百川Employee表的人（{emps.count()}人）:')
for e in emps:
    print(f'  {e.name:>8s} 身份证={e.id_card}')

# Excel里的所有人
print()
print('Excel里的所有人:')
for row_num in range(4, ws.max_row + 1):
    seq = ws.cell(row=row_num, column=1).value
    name = str(ws.cell(row=row_num, column=2).value or '').strip()
    id_card = str(ws.cell(row=row_num, column=3).value or '').strip()
    year_month_raw = str(ws.cell(row=row_num, column=4).value or '').strip()
    
    if name in {'小计', '合计', '在职人员', '退休人员', '家属统筹人员', ''}:
        continue
    if not id_card and not name:
        continue
    
    ym = f'{year_month_raw[:4]}-{year_month_raw[4:6]}' if len(year_month_raw) >= 6 else year_month_raw
    
    # Check if this person exists in Employee table
    emp = Employee.objects.filter(id_card=id_card).first()
    emp_status = '✅' if emp else '❌'
    emp_company = emp.company.name if emp and emp.company else 'N/A'
    
    # Check if record exists in DB
    if emp:
        db_rec = SocialRecord.objects.filter(employee=emp, year_month=ym).exists()
    else:
        db_rec = SocialRecord.objects.filter(company=bc, id_card=id_card, year_month=ym).exists()
    db_status = '✅' if db_rec else '❌'
    
    print(f'  {emp_status} {db_status} {name:>8s} {id_card} {ym} 员工={emp_company}')
