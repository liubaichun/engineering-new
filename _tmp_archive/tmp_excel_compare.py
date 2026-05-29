"""
对比 Excel 文件与数据库的社保记录
"""
import os, sys
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.insert(0, '/root/engineering-new')
import django
django.setup()

from openpyxl import load_workbook
from apps.finance.models import SocialRecord, Employee, Company

excel_path = '/root/.hermes/profiles/hermes-b001/cache/documents/doc_e2de73598919_深圳市百川软件科技发展有限公司_社保费申报明细_20260527.xlsx2401.xlsx'

wb = load_workbook(excel_path, data_only=True)
ws = wb.active

print(f'=== Excel 基本信息 ===')
print(f'Sheet名: {ws.title}')
print(f'行数: {ws.max_row}, 列数: {ws.max_column}')

# 打印前5行看表头
print()
print('=== 前5行原始数据 ===')
for row_num in range(1, min(6, ws.max_row + 1)):
    vals = []
    for col in range(ws.max_column):
        v = ws.cell(row=row_num, column=col+1).value
        vals.append(str(v)[:30] if v is not None else '')
    print(f'  Row{row_num}: {vals}')

# 扫描所有有效数据行（从Row4开始）
print()
print('=== Excel数据行扫描 ===')
valid_rows = 0
excel_records = []
skip_reasons = {}

for row_num in range(4, ws.max_row + 1):
    seq = ws.cell(row=row_num, column=1).value  # 序号
    name = str(ws.cell(row=row_num, column=2).value or '').strip()
    id_card = str(ws.cell(row=row_num, column=3).value or '').strip()
    year_month_raw = str(ws.cell(row=row_num, column=4).value or '').strip()
    
    # 跳过空行/小计行
    if name in {'小计', '合计', '在职人员', '退休人员', '家属统筹人员', ''}:
        reason = f'跳过(标签行): {name}'
        skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
        continue
    if not id_card and not name:
        skip_reasons['跳过(空行)'] = skip_reasons.get('跳过(空行)', 0) + 1
        continue
    if len(id_card) < 15:
        skip_reasons[f'跳过(身份证不合法): {id_card}'] = skip_reasons.get(f'跳过(身份证不合法): {id_card}', 0) + 1
        continue
    
    # 格式化年月
    if len(year_month_raw) >= 6:
        ym = f'{year_month_raw[:4]}-{year_month_raw[4:6]}'
    else:
        ym = year_month_raw
    
    # 获取关键金额
    total_receivable = ws.cell(row=row_num, column=6).value  # 应缴合计
    total_employee = ws.cell(row=row_num, column=7).value     # 个人部分
    total_company = ws.cell(row=row_num, column=8).value      # 单位部分
    
    valid_rows += 1
    excel_records.append({
        'row': row_num, 'name': name, 'id_card': id_card,
        'year_month': ym, 'total_company': total_company,
        'total_employee': total_employee, 'total_receivable': total_receivable
    })
    print(f'  Row{row_num}: {name:>8s} {id_card} {ym} 单位={total_company} 个人={total_employee}')

print(f'\nExcel有效数据行: {valid_rows}')
print(f'\n跳过统计:')
for reason, cnt in sorted(skip_reasons.items()):
    print(f'  {reason}: {cnt}行')

# 对比数据库
print()
print('=== 与数据库对比 ===')
bc = Company.objects.get(name__contains='百川')
db_records = SocialRecord.objects.filter(company=bc).order_by('year_month', 'employee__name')

print(f'数据库百川社保记录: {db_records.count()}条')

# 按 (name, year_month) 匹配
matched = 0
missing = []
for er in excel_records:
    # 查员工
    emp = Employee.objects.filter(id_card=er['id_card']).first()
    if emp:
        db_r = SocialRecord.objects.filter(employee=emp, year_month=er['year_month']).first()
        if db_r:
            matched += 1
        else:
            missing.append((er, emp, '员工匹配到但无此月份记录'))
    else:
        missing.append((er, None, '员工未匹配到Employee表'))

print(f'Excel有 DB也有: {matched}')
print(f'Excel有 DB没有: {len(missing)}')

print()
print('=== 缺失记录明细 ===')
for er, emp, reason in missing:
    print(f'  {er["name"]:>8s} {er["year_month"]} | {reason}')
    if emp:
        print(f'    → Employee id={emp.id}, company={emp.company.name if emp.company else "N/A"}')
    else:
        # 查这个身份证在哪个公司
        emp2 = Employee.objects.filter(id_card=er['id_card']).first()
        if emp2:
            print(f'    → 该身份证在系统里有Employee记录: {emp2.name}, 公司={emp2.company.name if emp2.company else "N/A"}')
        else:
            print(f'    → 该身份证在系统Employee表中不存在')
